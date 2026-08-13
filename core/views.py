import json
import random
from decimal import Decimal
from datetime import date, datetime, timedelta
from itertools import product

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

from . import views
from .decorators import permission_required
from .forms import VehicleLoadForm, SalesBillForm, EmployeeForm
from .models import (
    Category, Employee, Product, VehicleLoad, WarehouseStock, Vehicle, VehicleStock,
    SalesBill, SalesItem, Payment, Expense, UserProfile, Customer, 
    Cheque, OnlinePayment, MultiPayment, Bank, Supplier, Purchase, 
    PurchaseItem, StockMovement, StockTransfer, CreditCollection, DailySession
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.http import JsonResponse


@login_required
def dashboard(request):
    """Main Dashboard - Landing page after login"""
    
    today = date.today()
    start_of_month = date(today.year, today.month, 1)
    
    # ---------- TODAY'S DATA ----------
    today_bills = SalesBill.objects.filter(date=today)
    today_sales = today_bills.aggregate(total=Sum('net_total'))['total'] or Decimal('0')
    today_expense = Expense.objects.filter(date=today).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Today's Purchased (we'll use total sales as purchased since we don't have a purchase model)
    # You can replace this with actual purchase data if you have it
    today_purchased = Decimal('0')
    
    # Today's Profit (Sales - Expenses)
    today_profit = today_sales - today_expense
    
    # ---------- MONTHLY DATA ----------
    month_bills = SalesBill.objects.filter(date__gte=start_of_month, date__lte=today)
    month_sales = month_bills.aggregate(total=Sum('net_total'))['total'] or Decimal('0')
    month_expense = Expense.objects.filter(date__gte=start_of_month, date__lte=today).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    month_purchased = Decimal('0')  # Replace with actual purchase data if available
    
    # Monthly Profit
    month_profit = month_sales - month_expense
    
    # ---------- INVENTORY VALUE ----------
    warehouse_stocks = WarehouseStock.objects.all()
    inventory_value = Decimal('0')
    retail_value = Decimal('0')
    total_items = 0
    
    for stock in warehouse_stocks:
        product = stock.product
        inventory_value += stock.quantity * product.selling_price
        retail_value += stock.quantity * product.selling_price  # Same as inventory value for now
        total_items += 1
    
    # ---------- CREDIT BILLS ----------
    credit_bills = Payment.objects.filter(type='Credit')
    total_credit_bills = credit_bills.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # ---------- PURCHASE RETURNS (Monthly) ----------
    # Using Sales Returns as a proxy for purchase returns
    month_return_items = SalesItem.objects.filter(
        bill__in=month_bills,
        quantity__lt=0
    )
    month_purch_returns = month_return_items.aggregate(total=Sum('total'))['total'] or Decimal('0')
    month_purch_returns = abs(month_purch_returns)
    
    # ---------- TOTAL ITEMS ----------
    total_items = Product.objects.count()
    
    # ---------- RECENT PRODUCTS (for the table) ----------
    recent_products = Product.objects.all().order_by('-created_at')[:10]
    
    # ---------- CONTEXT ----------
    context = {
        'today': today,
        'today_sales': today_sales,
        'today_purchased': today_purchased,
        'today_expense': today_expense,
        'today_profit': today_profit,
        'month_sales': month_sales,
        'month_purchased': month_purchased,
        'month_expense': month_expense,
        'month_profit': month_profit,
        'inventory_value': inventory_value,
        'retail_value': retail_value,
        'total_credit_bills': total_credit_bills,
        'month_purch_returns': month_purch_returns,
        'total_items': total_items,
        'recent_products': recent_products,
        'start_of_month': start_of_month,
    }
    
    return render(request, 'core/dashboard.html', context)

    # ==================== SALES LIST ====================


@login_required
@permission_required('view_sales')
def sales_list(request):
    """List all sales with filters and actions"""
    
    bills = SalesBill.objects.all().select_related('vehicle', 'rep').order_by('-date', '-created_at')
    
    # ---------- FILTERS ----------
    search = request.GET.get('search', '')
    if search:
        bills = bills.filter(
            Q(invoice_no__icontains=search) | 
            Q(shop_name__icontains=search) |
            Q(shop_code__icontains=search)
        )
    
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if start_date:
        bills = bills.filter(date__gte=start_date)
    if end_date:
        bills = bills.filter(date__lte=end_date)
    
    vehicle_id = request.GET.get('vehicle', '')
    if vehicle_id and vehicle_id.isdigit():
        bills = bills.filter(vehicle_id=int(vehicle_id))
    
    vehicles = Vehicle.objects.filter(is_active=True)
    
    # Process each bill to get payment status and display info
    bill_list = []
    for bill in bills:
        # Calculate total paid (excluding Credit)
        paid_amount = bill.payments.exclude(type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Determine display status
        if paid_amount >= bill.net_total:
            display_status = 'COMPLETE'
        else:
            display_status = 'CREDIT'
        
        bill.display_status = display_status
        bill.display_customer = bill.shop_name or bill.shop_code or 'N/A'
        bill.total_items = bill.items.count()
        
        bill_list.append(bill)
    
    context = {
        'bills': bill_list,
        'vehicles': vehicles,
        'search': search,
        'start_date': start_date,
        'end_date': end_date,
        'selected_vehicle': vehicle_id,
        'total_count': len(bill_list),
    }
    
    return render(request, 'core/sales_list.html', context)

@login_required
@permission_required('view_sales')
def credit_list(request):
    """List all credit bills with outstanding balance > 0"""
    
    # Get all bills that have Credit payments
    credit_bill_ids = Payment.objects.filter(type='Credit').values_list('bill_id', flat=True).distinct()
    bills = SalesBill.objects.filter(id__in=credit_bill_ids).select_related('vehicle', 'rep').order_by('-date', '-created_at')
    
    # ---------- FILTERS ----------
    search = request.GET.get('search', '')
    if search:
        bills = bills.filter(
            Q(invoice_no__icontains=search) | 
            Q(shop_name__icontains=search) |
            Q(shop_code__icontains=search)
        )
    
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if start_date:
        bills = bills.filter(date__gte=start_date)
    if end_date:
        bills = bills.filter(date__lte=end_date)
    
    vehicle_id = request.GET.get('vehicle', '')
    if vehicle_id and vehicle_id.isdigit():
        bills = bills.filter(vehicle_id=int(vehicle_id))
    
    collection_status = request.GET.get('collection_status', '')
    rep_id = request.GET.get('rep', '')
    
    vehicles = Vehicle.objects.filter(is_active=True)
    reps = Employee.objects.filter(position='Rep', is_active=True)
    banks = Bank.objects.filter(is_active=True)
    
    # Process each bill to compute outstanding and collection info
    bill_list = []
    total_credit_outstanding = Decimal('0')
    pending_count = 0
    taken_count = 0
    collected_count = 0
    not_collected_count = 0
    
    for bill in bills:
        # Calculate total credit amount for this bill
        credit_amount = bill.payments.filter(type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        # Total paid (Cash + Cheque + Online) for this bill
        paid_amount = bill.payments.exclude(type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Determine credit status
        if paid_amount >= bill.net_total:
            credit_status = 'PAID'
        elif paid_amount > 0:
            credit_status = 'PARTIAL'
        else:
            credit_status = 'OUTSTANDING'
        
        outstanding = bill.net_total - paid_amount
        
        # If outstanding <= 0, skip this bill (fully paid, no credit)
        if outstanding <= 0:
            continue
        
        total_credit_outstanding += outstanding
        
        # Get or create collection record
        try:
            collection, created = CreditCollection.objects.get_or_create(
                sales_bill=bill,
                defaults={'status': 'PENDING'}
            )
        except:
            collection = None
        
        # Apply collection status filter
        if collection_status and collection and collection_status != collection.status:
            continue
        
        # Apply rep filter
        if rep_id and rep_id.isdigit() and collection and collection.rep_id != int(rep_id):
            continue
        
        # Update counts
        if collection:
            if collection.status == 'PENDING':
                pending_count += 1
            elif collection.status == 'TAKEN':
                taken_count += 1
            elif collection.status == 'COLLECTED':
                collected_count += 1
            elif collection.status == 'NOT_COLLECTED':
                not_collected_count += 1
        
        display_customer = bill.shop_name or bill.shop_code or 'N/A'
        
        # ✅ The SalesBill object is stored as 'bill'
        bill_list.append({
            'bill': bill,
            'credit_amount': credit_amount,
            'paid_amount': paid_amount,
            'outstanding': outstanding,
            'credit_status': credit_status,
            'customer': display_customer,
            'total_items': bill.items.count(),
            'collection': collection,
        })
    
    status_choices = CreditCollection.STATUS_CHOICES if hasattr(CreditCollection, 'STATUS_CHOICES') else []
    
    context = {
        'bills': bill_list,
        'vehicles': vehicles,
        'reps': reps,
        'banks': banks,
        'search': search,
        'start_date': start_date,
        'end_date': end_date,
        'selected_vehicle': vehicle_id,
        'selected_collection_status': collection_status,
        'selected_rep': rep_id,
        'total_credit_outstanding': total_credit_outstanding,
        'total_count': len(bill_list),
        'pending_count': pending_count,
        'taken_count': taken_count,
        'collected_count': collected_count,
        'not_collected_count': not_collected_count,
        'status_choices': status_choices,
    }
    
    return render(request, 'core/credit_list.html', context)


@login_required
@permission_required('create_sales')
def credit_collection_take(request):
    """Take credit bills for collection (morning)"""
    if request.method == 'POST':
        rep_id = request.POST.get('rep')
        bill_ids = request.POST.getlist('bill_ids')
        collection_date = request.POST.get('collection_date', date.today())
        
        if not rep_id:
            messages.error(request, 'Please select a rep.')
            return redirect('core:credit_list')
        
        if not bill_ids:
            messages.error(request, 'Please select at least one bill.')
            return redirect('core:credit_list')
        
        rep = get_object_or_404(Employee, id=rep_id)
        
        taken_count = 0
        for bill_id in bill_ids:
            bill = get_object_or_404(SalesBill, id=bill_id)
            
            collection, created = CreditCollection.objects.get_or_create(
                sales_bill=bill,
                defaults={
                    'rep': rep,
                    'date_taken': collection_date,
                    'status': 'TAKEN',
                }
            )
            
            if not created:
                collection.rep = rep
                collection.date_taken = collection_date
                collection.status = 'TAKEN'
                collection.save()
            taken_count += 1
        
        messages.success(request, f'Successfully taken {taken_count} credit bill(s) for collection.')
        return redirect('core:credit_list')
    
    return redirect('core:credit_list')


@login_required
@permission_required('create_sales')
def credit_collection_return(request):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('core:credit_list')
    
    collection_id = request.POST.get('collection_id')
    action = request.POST.get('action')
    notes = request.POST.get('notes', '')
    
    if not collection_id or not action:
        messages.error(request, 'Missing required fields.')
        return redirect('core:credit_list')
    
    collection = get_object_or_404(CreditCollection, id=collection_id)
    
    if action == 'collected':
        collection.status = 'COLLECTED'
        collection.date_collected = date.today()
        collection.notes = notes
        # No payment processing here – payment is handled separately via Pay Now
    elif action == 'not_collected':
        reason = request.POST.get('reason')
        if not reason:
            messages.error(request, 'Please select a reason for not collected.')
            return redirect('core:credit_list')
        collection.status = 'NOT_COLLECTED'
        collection.not_collected_reason = reason
        collection.notes = notes
    else:
        messages.error(request, 'Invalid action.')
        return redirect('core:credit_list')
    
    collection.save()
    messages.success(request, f'Bill {collection.sales_bill.invoice_no} marked as {collection.get_status_display()}.')
    return redirect('core:credit_list')


@login_required
@permission_required('view_sales')
def credit_collection_report(request):
    """Collection report by date and rep"""
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    rep_id = request.GET.get('rep', '')
    
    collections = CreditCollection.objects.filter(
        date_taken__gte=start_date,
        date_taken__lte=end_date
    ).select_related('sales_bill', 'rep', 'collected_by')
    
    if rep_id and rep_id.isdigit():
        collections = collections.filter(rep_id=int(rep_id))
    
    # Statistics
    total_taken = collections.count()
    total_collected = collections.filter(status='COLLECTED').count()
    total_not_collected = collections.filter(status='NOT_COLLECTED').count()
    total_pending = collections.filter(status='PENDING').count()
    
    collected_amount = collections.filter(status='COLLECTED').aggregate(total=Sum('collection_amount'))['total'] or 0
    
    # Rep breakdown
    rep_breakdown = collections.values('rep__name').annotate(
        taken=Count('id'),
        collected=Count('id', filter=Q(status='COLLECTED')),
        not_collected=Count('id', filter=Q(status='NOT_COLLECTED')),
        amount=Sum('collection_amount')
    )
    
    reps = Employee.objects.filter(position='Rep', is_active=True)
    
    context = {
        'collections': collections,
        'start_date': start_date,
        'end_date': end_date,
        'selected_rep': rep_id,
        'reps': reps,
        'total_taken': total_taken,
        'total_collected': total_collected,
        'total_not_collected': total_not_collected,
        'total_pending': total_pending,
        'collected_amount': collected_amount,
        'rep_breakdown': rep_breakdown,
        'today': today,
    }
    return render(request, 'core/credit_collection_report.html', context)

@login_required
@permission_required('create_sales')
def credit_collection_retake(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})
    
    collection_id = request.POST.get('collection_id')
    if not collection_id:
        return JsonResponse({'success': False, 'message': 'Collection ID is required.'})
    
    collection = get_object_or_404(CreditCollection, id=collection_id)
    
    # Only allow retake if status is NOT_COLLECTED
    if collection.status != 'NOT_COLLECTED':
        return JsonResponse({'success': False, 'message': 'Bill is not in NOT_COLLECTED status.'})
    
    from datetime import date
    # Move back to PENDING so it can be taken again
    collection.status = 'PENDING'
    collection.date_taken = None          # Clear taken date
    collection.not_collected_reason = None
    collection.save()
    
    return JsonResponse({'success': True, 'message': 'Bill moved back to PENDING.'})


@login_required
@permission_required('create_sales')
def mark_credit_paid(request, bill_id):
    """Mark a credit bill as fully paid by adding a Cash payment"""
    bill = get_object_or_404(SalesBill, id=bill_id)
    
    # Check if there's any outstanding amount
    paid_total = bill.payments.exclude(type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
    outstanding = bill.net_total - paid_total
    
    if outstanding <= 0:
        messages.warning(request, 'This bill is already fully paid.')
        return redirect('core:credit_list')
    
    if request.method == 'POST':
        # Add a Cash payment for the outstanding amount
        Payment.objects.create(
            bill=bill,
            type='Cash',
            amount=outstanding
        )
        messages.success(request, f'Credit bill {bill.invoice_no} marked as PAID.')
        return redirect('core:credit_list')
    
    # GET - show confirmation page
    context = {
    'bill': bill,
    'outstanding': outstanding,
    'customer': bill.shop_name or bill.shop_code or 'N/A',
    'paid_amount': bill.net_total - outstanding,  # Add this line
}
    
    return render(request, 'core/credit_pay.html', context)


@login_required
@permission_required('view_sales')
def sales_detail(request, bill_id):
    """View individual sale details"""
    
    bill = get_object_or_404(SalesBill, id=bill_id)
    items = bill.items.all().select_related('product')
    payments = bill.payments.all()
    
    # Get status
    if payments.exists():
        payment_types = [p.type for p in payments]
        if 'Credit' in payment_types and 'Cash' not in payment_types and 'Cheque' not in payment_types:
            status = 'CREDIT'
        elif 'Cash' in payment_types or 'Cheque' in payment_types:
            status = 'COMPLETE'
        else:
            status = 'PENDING'
    else:
        status = 'PENDING'
    
    context = {
        'bill': bill,
        'items': items,
        'payments': payments,
        'status': status,
    }
    
    return render(request, 'core/sales_detail.html', context)


def login_view(request):
    """User Login Page"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Auto-create profile if missing
            if not hasattr(user, 'profile'):
                UserProfile.objects.create(
                    user=user,
                    role='Viewer',
                    phone='',
                    is_active=True
                )
                messages.info(request, f'Welcome {user.username}! Your profile has been created.')
            
            login(request, user)
            messages.success(request, f'Welcome back, {user.username}!')
            return redirect('/')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'core/login.html')


def logout_view(request):
    """User Logout"""
    logout(request)
    messages.success(request, 'You have been logged out.')
    return redirect('/login/')  # ✅ FIXED


def register_view(request):
    """Create a new user account (Admin only)"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        role = request.POST.get('role', 'Viewer')
        phone = request.POST.get('phone', '')
        
        if password != password2:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'core/register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'core/register.html')
        
        user = User.objects.create_user(username=username, password=password)
        UserProfile.objects.create(user=user, role=role, phone=phone)
        messages.success(request, f'User {username} created successfully!')
        return redirect('/admin-dashboard/')  # ✅ FIXED
    
    return render(request, 'core/register.html')


@login_required
def admin_dashboard(request):
    """Admin panel to manage users"""
    try:
        profile = request.user.profile
        if profile.role != 'Admin':
            messages.error(request, 'Only Admins can access user management.')
            return redirect('/reports/')  # ✅ FIXED
    except UserProfile.DoesNotExist:
        messages.error(request, 'Profile not found.')
        return redirect('/logout/')  # ✅ FIXED
    
    users = User.objects.all().select_related('profile')
    return render(request, 'core/admin_dashboard.html', {'users': users})


@login_required
def toggle_user_status(request, user_id):
    """Enable/Disable a user account (Admin only)"""
    try:
        profile = request.user.profile
        if profile.role != 'Admin':
            messages.error(request, 'Only Admins can manage users.')
            return redirect('/reports/')  # ✅ FIXED
    except UserProfile.DoesNotExist:
        messages.error(request, 'Profile not found.')
        return redirect('/logout/')  # ✅ FIXED
    
    target_user = get_object_or_404(User, id=user_id)
    if target_user == request.user:
        messages.error(request, 'You cannot disable your own account.')
        return redirect('/admin-dashboard/')  # ✅ FIXED
    
    target_user.is_active = not target_user.is_active
    target_user.save()
    status = 'enabled' if target_user.is_active else 'disabled'
    messages.success(request, f'User {target_user.username} has been {status}.')
    return redirect('/admin-dashboard/')  # ✅ FIXED


@login_required
@permission_required('load_vehicle')
def load_vehicle(request):
    products = Product.objects.all()
    stock_dict = {stock.product_id: stock for stock in WarehouseStock.objects.all()}
    
    product_list = []
    for p in products:
        stock_obj = stock_dict.get(p.id)
        product_list.append({
            'product': p,
            'current_stock': stock_obj.quantity if stock_obj else 0
        })

    if request.method == 'POST':
        form = VehicleLoadForm(request.POST)
        if form.is_valid():
            vehicle = form.cleaned_data['vehicle']
            
            with transaction.atomic():
                for item in product_list:
                    product = item['product']
                    quantity_key = f'qty_{product.id}'
                    quantity_str = request.POST.get(quantity_key, '0')
                    
                    try:
                        quantity = Decimal(quantity_str)
                    except ValueError:
                        quantity = Decimal('0')
                    
                    if quantity > 0:
                        warehouse_stock = WarehouseStock.objects.get(product=product)
                        if warehouse_stock.quantity < quantity:
                            messages.error(request, f'❌ Not enough stock for {product.name}. Available: {warehouse_stock.quantity}')
                            return render(request, 'core/load_vehicle.html', {
                                'form': form, 
                                'product_list': product_list
                            })
                        
                        warehouse_stock.quantity -= quantity
                        warehouse_stock.save()
                        
                        vehicle_stock, created = VehicleStock.objects.get_or_create(
                            vehicle=vehicle, 
                            product=product
                        )
                        vehicle_stock.quantity += quantity
                        vehicle_stock.save()
                
                messages.success(request, f'✅ Successfully loaded stock to {vehicle.vehicle_number} - {vehicle.driver_name}')
                return redirect('/load/')  # ✅ FIXED (was 'load_vehicle')
    else:
        form = VehicleLoadForm()

        VehicleLoad.objects.create(
    vehicle=vehicle,
    product=product,
    quantity=quantity,
    loaded_by=request.user,
    notes=f"Loaded from warehouse"
)

    return render(request, 'core/load_vehicle.html', {
        'form': form,
        'product_list': product_list

        
    })


@login_required
@permission_required('view_reports')
def reports_dashboard(request):
    today = datetime.today().date()
    start_date_str = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        start_date = today
        end_date = today
    
    bills = SalesBill.objects.filter(date__range=[start_date, end_date])
    bill_ids = bills.values_list('id', flat=True)
    
    total_sales = bills.aggregate(total=Sum('net_total'))['total'] or Decimal('0')
    cash_total = Payment.objects.filter(bill__in=bills, type='Cash').aggregate(total=Sum('amount'))['total'] or Decimal('0')
    credit_total = Payment.objects.filter(bill__in=bills, type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
    cheque_total = Payment.objects.filter(bill__in=bills, type='Cheque').aggregate(total=Sum('amount'))['total'] or Decimal('0')
    discount_total = bills.aggregate(total=Sum('discount_total'))['total'] or Decimal('0')
    
    return_items = SalesItem.objects.filter(bill__in=bills, quantity__lt=0)
    return_total = return_items.aggregate(total=Sum('total'))['total'] or Decimal('0')
    return_total_abs = abs(return_total)
    
    expenses = Expense.objects.filter(date__range=[start_date, end_date])
    expense_total = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    net_profit = total_sales - return_total_abs - expense_total
    
    vehicles = Vehicle.objects.filter(is_active=True)
    vehicle_data = []
    for vehicle in vehicles:
        vehicle_bills = bills.filter(vehicle=vehicle)
        if vehicle_bills.exists():
            veh_sales = vehicle_bills.aggregate(total=Sum('net_total'))['total'] or Decimal('0')
            veh_cash = Payment.objects.filter(bill__in=vehicle_bills, type='Cash').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            veh_credit = Payment.objects.filter(bill__in=vehicle_bills, type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            veh_cheque = Payment.objects.filter(bill__in=vehicle_bills, type='Cheque').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            veh_discount = vehicle_bills.aggregate(total=Sum('discount_total'))['total'] or Decimal('0')
            veh_returns = SalesItem.objects.filter(bill__in=vehicle_bills, quantity__lt=0).aggregate(total=Sum('total'))['total'] or Decimal('0')
            veh_returns_abs = abs(veh_returns)
            veh_expenses = Expense.objects.filter(vehicle=vehicle, date__range=[start_date, end_date]).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            
            vehicle_data.append({
                'vehicle': vehicle,
                'sales': veh_sales,
                'cash': veh_cash,
                'credit': veh_credit,
                'cheque': veh_cheque,
                'discount': veh_discount,
                'returns': veh_returns_abs,
                'expenses': veh_expenses,
                'net': veh_sales - veh_returns_abs - veh_expenses,
            })
    
    return_products = SalesItem.objects.filter(
        bill__in=bills, 
        quantity__lt=0
    ).values('product__name', 'product__code').annotate(
        total_returned_qty=Sum('quantity'),
        total_returned_value=Sum('total')
    ).order_by('total_returned_value')
    
    expense_by_category = expenses.values('category').annotate(total=Sum('amount')).order_by('-total')
    
    daily_data = []
    if (end_date - start_date).days > 1:
        current_date = start_date
        while current_date <= end_date:
            day_bills = SalesBill.objects.filter(date=current_date)
            if day_bills.exists():
                day_sales = day_bills.aggregate(total=Sum('net_total'))['total'] or Decimal('0')
                day_cash = Payment.objects.filter(bill__in=day_bills, type='Cash').aggregate(total=Sum('amount'))['total'] or Decimal('0')
                day_credit = Payment.objects.filter(bill__in=day_bills, type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
                day_cheque = Payment.objects.filter(bill__in=day_bills, type='Cheque').aggregate(total=Sum('amount'))['total'] or Decimal('0')
                day_expenses = Expense.objects.filter(date=current_date).aggregate(total=Sum('amount'))['total'] or Decimal('0')
                daily_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'sales': day_sales,
                    'cash': day_cash,
                    'credit': day_credit,
                    'cheque': day_cheque,
                    'expenses': day_expenses,
                })
            current_date += timedelta(days=1)
    
    context = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_sales': total_sales,
        'cash_total': cash_total,
        'credit_total': credit_total,
        'cheque_total': cheque_total,
        'discount_total': discount_total,
        'return_total': return_total_abs,
        'expense_total': expense_total,
        'net_profit': net_profit,
        'vehicle_data': vehicle_data,
        'return_products': return_products,
        'expense_by_category': expense_by_category,
        'daily_data': daily_data,
        'bill_count': bills.count(),
        'today': today,
    }
    
    return render(request, 'core/reports.html', context)


@login_required
@permission_required('view_employees')  # We'll add this permission
def employee_list(request):
    """List all employees with filters"""
    employees = Employee.objects.all()
    
    # Search
    search = request.GET.get('search', '')
    if search:
        employees = employees.filter(
            Q(name__icontains=search) | 
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Filter by position
    position = request.GET.get('position', '')
    if position:
        employees = employees.filter(position=position)
    
    # Filter by active status
    status = request.GET.get('status', '')
    if status == 'active':
        employees = employees.filter(is_active=True)
    elif status == 'inactive':
        employees = employees.filter(is_active=False)
    
    context = {
        'employees': employees,
        'search': search,
        'selected_position': position,
        'selected_status': status,
        'positions': Employee.POSITION_CHOICES,
        'total_count': employees.count(),
    }
    return render(request, 'core/employee_list.html', context)


@login_required
@permission_required('manage_employees')
def employee_add(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            employee = form.save(commit=False)
            # Ensure rep_code is None if position not Rep
            if employee.position != 'Rep':
                employee.rep_code = None
                employee.rep_invoice_counter = 0
            employee.save()
            messages.success(request, f'Employee {employee.name} added successfully!')
            return redirect('core:employee_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = EmployeeForm()
    
    return render(request, 'core/employee_form.html', {
        'form': form,
        'action': 'Add',
        'employee': None, 
        'positions': Employee.POSITION_CHOICES, 
        'today': date.today(),
    })


@login_required
@permission_required('manage_employees')
def employee_edit(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            emp = form.save(commit=False)
            if emp.position != 'Rep':
                emp.rep_code = None
                emp.rep_invoice_counter = 0
            emp.save()
            messages.success(request, f'Employee {employee.name} updated successfully!')
            return redirect('core:employee_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = EmployeeForm(instance=employee)
    
    return render(request, 'core/employee_form.html', {
        'form': form,
        'action': 'Edit',
        'employee': employee,
        'positions': Employee.POSITION_CHOICES,
        'today': date.today(),
    })


@login_required
@permission_required('manage_employees')
def employee_toggle_status(request, employee_id):
    """Activate or deactivate an employee"""
    employee = get_object_or_404(Employee, id=employee_id)
    employee.is_active = not employee.is_active
    employee.save()
    status = 'activated' if employee.is_active else 'deactivated'
    messages.success(request, f'Employee {employee.name} has been {status}.')
    return redirect('core:employee_list')


@login_required
@permission_required('manage_employees')
def employee_delete(request, employee_id):
    """Permanently delete an employee (use with caution)"""
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, f'Employee {employee.name} deleted permanently.')
        return redirect('core:employee_list')
    
    # GET request - show confirmation
    return render(request, 'core/employee_confirm_delete.html', {'employee': employee})


@csrf_exempt
def add_bank_api(request):
    """API endpoint to add a new bank"""
    if request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Bank name is required.'})
        
        if Bank.objects.filter(name__iexact=name).exists():
            return JsonResponse({'success': False, 'error': 'Bank already exists.'})
        
        bank = Bank.objects.create(name=name)
        return JsonResponse({'success': True, 'id': bank.id, 'name': bank.name})
    
    return JsonResponse({'success': False, 'error': 'Invalid request.'})


@login_required
@permission_required('view_products')
def product_list(request):
    """List all products with search and filter"""
    products = Product.objects.all().select_related('category')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        products = products.filter(
            Q(name__icontains=search) | 
            Q(code__icontains=search)
        )
    
    # Filter by category
    category_id = request.GET.get('category', '')
    if category_id and category_id.isdigit():
        products = products.filter(category_id=int(category_id))
    
    # Filter by status
    status = request.GET.get('status', '')
    if status == 'active':
        products = products.filter(is_active=True)
    elif status == 'inactive':
        products = products.filter(is_active=False)
    
    categories = Category.objects.filter(is_active=True)
    
    context = {
        'products': products,
        'categories': categories,
        'search': search,
        'selected_category': category_id,
        'selected_status': status,
        'total_count': products.count(),
    }
    return render(request, 'core/product_list.html', context)


@login_required
@permission_required('manage_products')
def product_add(request):
    """Add a new product"""
    categories = Category.objects.filter(is_active=True)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        category_id = request.POST.get('category')
        unit = request.POST.get('unit')
        selling_price = request.POST.get('selling_price')
        cost_price = request.POST.get('cost_price')
        notes = request.POST.get('notes', '')
        is_active = request.POST.get('is_active') == 'on'
        
        # Validate
        if not name or not code or not selling_price:
            messages.error(request, 'Name, Code, and Selling Price are required.')
            return render(request, 'core/product_form.html', {
                'categories': categories,
                'product': None,
                'action': 'Add',
            })
        
        if Product.objects.filter(code=code).exists():
            messages.error(request, f'Product with code "{code}" already exists.')
            return render(request, 'core/product_form.html', {
                'categories': categories,
                'product': None,
                'action': 'Add',
            })
        
        # Create product
        product = Product.objects.create(
            name=name,
            code=code,
            category_id=category_id if category_id else None,
            unit=unit or 'Pcs',
            selling_price=selling_price,
            cost_price=cost_price if cost_price else None,
            notes=notes,
            is_active=is_active
        )
        
        # Create warehouse stock entry
        WarehouseStock.objects.create(product=product, quantity=0)
        
        messages.success(request, f'Product "{product.name}" added successfully!')
        return redirect('core:product_list')
    
    return render(request, 'core/product_form.html', {
        'categories': categories,
        'product': None,
        'action': 'Add',
        'units': Product.UNIT_CHOICES,
    })


@login_required
@permission_required('manage_products')
def product_edit(request, product_id):
    """Edit an existing product"""
    product = get_object_or_404(Product, id=product_id)
    categories = Category.objects.filter(is_active=True)
    
    if request.method == 'POST':
        product.name = request.POST.get('name')
        product.code = request.POST.get('code')
        product.category_id = request.POST.get('category') or None
        product.unit = request.POST.get('unit') or 'Pcs'
        product.selling_price = request.POST.get('selling_price')
        product.cost_price = request.POST.get('cost_price') or None
        product.notes = request.POST.get('notes', '')
        product.is_active = request.POST.get('is_active') == 'on'
        product.save()
        
        messages.success(request, f'Product "{product.name}" updated successfully!')
        return redirect('core:product_list')
    
    return render(request, 'core/product_form.html', {
        'product': product,
        'categories': categories,
        'action': 'Edit',
        'units': Product.UNIT_CHOICES,
    })


@login_required
@permission_required('manage_products')
def product_delete(request, product_id):
    """Delete a product (with confirmation)"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        product.delete()
        messages.success(request, f'Product "{product.name}" deleted successfully.')
        return redirect('core:product_list')
    
    return render(request, 'core/product_confirm_delete.html', {'product': product})


@login_required
@permission_required('manage_products')
def product_toggle_status(request, product_id):
    """Activate or deactivate a product"""
    product = get_object_or_404(Product, id=product_id)
    product.is_active = not product.is_active
    product.save()
    status = 'activated' if product.is_active else 'deactivated'
    messages.success(request, f'Product "{product.name}" has been {status}.')
    return redirect('core:product_list')


@login_required
@permission_required('manage_products')
def category_list(request):
    """List all categories"""
    categories = Category.objects.all().annotate(product_count=Count('products'))
    
    search = request.GET.get('search', '')
    if search:
        categories = categories.filter(name__icontains=search)
    
    context = {
        'categories': categories,
        'search': search,
        'total_count': categories.count(),
    }
    return render(request, 'core/category_list.html', context)


@login_required
@permission_required('manage_products')
def category_add(request):
    """Add a new category"""
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        
        if not name:
            messages.error(request, 'Category name is required.')
            return render(request, 'core/category_form.html', {'action': 'Add', 'category': None})
        
        if Category.objects.filter(name__iexact=name).exists():
            messages.error(request, f'Category "{name}" already exists.')
            return render(request, 'core/category_form.html', {'action': 'Add', 'category': None})
        
        Category.objects.create(name=name, description=description)
        messages.success(request, f'Category "{name}" added successfully!')
        return redirect('core:category_list')
    
    return render(request, 'core/category_form.html', {'action': 'Add', 'category': None})


@login_required
@permission_required('manage_products')
def category_edit(request, category_id):
    """Edit an existing category"""
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        
        if not name:
            messages.error(request, 'Category name is required.')
            return render(request, 'core/category_form.html', {'action': 'Edit', 'category': category})
        
        if Category.objects.filter(name__iexact=name).exclude(id=category_id).exists():
            messages.error(request, f'Category "{name}" already exists.')
            return render(request, 'core/category_form.html', {'action': 'Edit', 'category': category})
        
        category.name = name
        category.description = description
        category.save()
        
        messages.success(request, f'Category "{category.name}" updated successfully!')
        return redirect('core:category_list')
    
    return render(request, 'core/category_form.html', {'action': 'Edit', 'category': category})


@login_required
@permission_required('manage_products')
def category_delete(request, category_id):
    """Delete a category (with confirmation)"""
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        category.delete()
        messages.success(request, f'Category "{category.name}" deleted successfully.')
        return redirect('core:category_list')
    
    return render(request, 'core/category_confirm_delete.html', {'category': category})


@login_required
@permission_required('view_products')
def stock_view(request):
    """View all stock (warehouse + vehicles)"""
    products = Product.objects.filter(is_active=True).select_related('category')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        products = products.filter(
            Q(name__icontains=search) | 
            Q(code__icontains=search)
        )
    
    # Filter by category
    category_id = request.GET.get('category', '')
    if category_id and category_id.isdigit():
        products = products.filter(category_id=int(category_id))
    
    # Prepare stock data
    stock_data = []
    for product in products:
        warehouse_stock = product.get_warehouse_stock()
        vehicle_stock = product.get_vehicle_stock()
        total_stock = warehouse_stock + vehicle_stock
        
        # Get vehicle stock details
        vehicle_stocks = VehicleStock.objects.filter(product=product).select_related('vehicle')
        vehicle_details = []
        for vs in vehicle_stocks:
            vehicle_details.append({
                'vehicle': vs.vehicle.vehicle_number,
                'driver': vs.vehicle.driver_name,
                'quantity': vs.quantity,
            })
        
        stock_data.append({
            'product': product,
            'warehouse_stock': warehouse_stock,
            'vehicle_stock': vehicle_stock,
            'total_stock': total_stock,
            'vehicle_details': vehicle_details,
            'profit_margin': product.get_profit_margin(),
        })
    
    # Calculate totals
    total_warehouse = sum(item['warehouse_stock'] for item in stock_data)
    total_vehicle = sum(item['vehicle_stock'] for item in stock_data)
    total_stock_all = sum(item['total_stock'] for item in stock_data)
    
    categories = Category.objects.filter(is_active=True)
    
    context = {
        'stock_data': stock_data,
        'categories': categories,
        'search': search,
        'selected_category': category_id,
        'total_warehouse': total_warehouse,
        'total_vehicle': total_vehicle,
        'total_stock_all': total_stock_all,
        'total_products': len(stock_data),
    }
    return render(request, 'core/stock_view.html', context)


@login_required
@permission_required('view_purchases')
def purchase_list(request):
    """List all purchases with filters"""
    purchases = Purchase.objects.all().select_related('supplier', 'rep')
    
    # Search by Invoice No or PO Number
    search = request.GET.get('search', '')
    if search:
        purchases = purchases.filter(
            Q(invoice_no__icontains=search) | 
            Q(po_number__icontains=search) |
            Q(supplier__name__icontains=search)
        )
    
    # Date filter
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if start_date:
        purchases = purchases.filter(purchase_date__gte=start_date)
    if end_date:
        purchases = purchases.filter(purchase_date__lte=end_date)
    
    # Supplier filter
    supplier_id = request.GET.get('supplier', '')
    if supplier_id and supplier_id.isdigit():
        purchases = purchases.filter(supplier_id=int(supplier_id))
    
    # Payment Status filter
    payment_status = request.GET.get('payment_status', '')
    if payment_status:
        purchases = purchases.filter(payment_status=payment_status)
    
    # Status filter
    status = request.GET.get('status', '')
    if status:
        purchases = purchases.filter(status=status)
    
    suppliers = Supplier.objects.filter(is_active=True)
    
    # Calculate totals
    total_cost_sum = purchases.aggregate(total=Sum('total'))['total'] or 0
    
    context = {
        'purchases': purchases,
        'suppliers': suppliers,
        'search': search,
        'start_date': start_date,
        'end_date': end_date,
        'selected_supplier': supplier_id,
        'selected_payment_status': payment_status,
        'selected_status': status,
        'total_count': purchases.count(),
        'total_cost_sum': total_cost_sum,
        'payment_status_choices': Purchase.PAYMENT_STATUS_CHOICES,
        'status_choices': Purchase.STATUS_CHOICES,
    }
    return render(request, 'core/purchase_list.html', context)


@login_required
@permission_required('manage_purchases')
def purchase_add(request):
    """Add a new purchase with FOC support"""
    suppliers = Supplier.objects.filter(is_active=True)
    reps = Employee.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True)
    
    if request.method == 'POST':
        supplier_id = request.POST.get('supplier')
        rep_id = request.POST.get('rep')
        po_number = request.POST.get('po_number')
        invoice_no = request.POST.get('invoice_no')
        purchase_date = request.POST.get('purchase_date')
        due_date = request.POST.get('due_date')
        tax_rate = int(request.POST.get('tax_rate', 0))
        tax_invoice_no = request.POST.get('tax_invoice_no', '')
        paid_amount = Decimal(request.POST.get('paid_amount', '0') or '0')
        notes = request.POST.get('notes', '')
        payment_status = request.POST.get('payment_status', 'PENDING')
        
        # Validate
        if not supplier_id or not invoice_no:
            messages.error(request, 'Supplier and Invoice Number are required.')
            return render(request, 'core/purchase_add.html', {
                'suppliers': suppliers,
                'reps': reps,
                'products': products,
                'tax_rates': Purchase.TAX_RATE_CHOICES,
                'today': date.today(),
            })
        
        supplier = get_object_or_404(Supplier, id=supplier_id)
        rep = None
        if rep_id:
            rep = get_object_or_404(Employee, id=rep_id)
        
        # Create purchase
        purchase = Purchase.objects.create(
            supplier=supplier,
            rep=rep,
            po_number=po_number or '',
            invoice_no=invoice_no,
            purchase_date=purchase_date or date.today(),
            due_date=due_date if due_date else None,
            tax_rate=tax_rate,
            tax_invoice_no=tax_invoice_no,
            paid_amount=paid_amount,
            payment_status=payment_status,
            notes=notes,
            status='PENDING',
            created_by=request.user,
        )
        
        # Process items
        subtotal = Decimal('0')
        foc_value = Decimal('0')
        items_data = []
        
        item_count = int(request.POST.get('item_count', 0))
        for i in range(item_count):
            product_id = request.POST.get(f'product_{i}')
            quantity = Decimal(request.POST.get(f'qty_{i}', '0') or '0')
            cost_price = Decimal(request.POST.get(f'cost_{i}', '0') or '0')
            retail_price = Decimal(request.POST.get(f'retail_{i}', '0') or '0')
            wholesale_price = Decimal(request.POST.get(f'wholesale_{i}', '0') or '0')
            is_foc = request.POST.get(f'is_foc_{i}', 'false') == 'true'
            
            if not product_id or quantity <= 0 or cost_price <= 0:
                continue
            
            product = get_object_or_404(Product, id=product_id)
            total = quantity * cost_price
            retail_total = quantity * retail_price
            wholesale_total = quantity * wholesale_price
            
            PurchaseItem.objects.create(
                purchase=purchase,
                product=product,
                quantity=quantity,
                cost_price=cost_price,
                retail_price=retail_price,
                wholesale_price=wholesale_price,
                total=total,
                retail_total=retail_total,
                wholesale_total=wholesale_total,
                is_foc=is_foc,
            )
            
            # Subtotal includes ALL items (regular + FOC) for valuation
            subtotal += total
            # FOC value tracked separately (but same as total for FOC items)
            if is_foc:
                foc_value += total
            
            items_data.append({
                'product': product,
                'quantity': quantity,
                'cost_price': cost_price,
                'retail_price': retail_price,
                'wholesale_price': wholesale_price,
                'total': total,
                'is_foc': is_foc,
            })
        
        if items_data:
            # Update purchase totals
            purchase.subtotal = subtotal
            purchase.tax_amount = (subtotal * tax_rate) / 100
            purchase.total = subtotal + purchase.tax_amount
            purchase.save()
            
            messages.success(request, f'Purchase #{purchase.invoice_no} created successfully!')
            return redirect('core:purchase_list')
        else:
            purchase.delete()
            messages.error(request, 'Please add at least one item to the purchase.')
            return render(request, 'core/purchase_add.html', {
                'suppliers': suppliers,
                'reps': reps,
                'products': products,
                'tax_rates': Purchase.TAX_RATE_CHOICES,
                'today': date.today(),
            })
    
    # GET request
    return render(request, 'core/purchase_add.html', {
        'suppliers': suppliers,
        'reps': reps,
        'products': products,
        'tax_rates': Purchase.TAX_RATE_CHOICES,
        'today': date.today(),
        'random_po': f"{random.randint(100000, 999999)}",
    })


@login_required
@permission_required('manage_purchases')
def purchase_detail(request, purchase_id):
    """View purchase details"""
    purchase = get_object_or_404(Purchase, id=purchase_id)
    items = purchase.items.all().select_related('product')
    
    context = {
        'purchase': purchase,
        'items': items,
    }
    return render(request, 'core/purchase_detail.html', context)


@login_required
@permission_required('manage_purchases')
def purchase_receive(request, purchase_id):
    """Mark purchase as received - updates stock"""
    purchase = get_object_or_404(Purchase, id=purchase_id)
    items = purchase.items.all().select_related('product')
    
    if request.method == 'POST':
        with transaction.atomic():
            for item in items:
                # Update product cost price (if changed)
                if item.cost_price != item.product.cost_price:
                    item.product.cost_price = item.cost_price
                    item.product.save()
                
                # Get current warehouse stock
                warehouse_stock, created = WarehouseStock.objects.get_or_create(
                    product=item.product
                )
                
                old_stock = warehouse_stock.quantity
                
                # Add quantity to warehouse stock
                warehouse_stock.quantity += item.quantity
                warehouse_stock.save()
                
                # Create stock movement log
                StockMovement.objects.create(
                    product=item.product,
                    movement_type='PURCHASE',
                    quantity=item.quantity,
                    previous_stock=old_stock,
                    new_stock=warehouse_stock.quantity,
                    reference=f"PO: {purchase.invoice_no}",
                    notes=f"Purchase received from {purchase.supplier.name}",
                    created_by=request.user,
                )
            
            # Update purchase status
            purchase.status = 'RECEIVED'
            purchase.received_at = timezone.now()
            purchase.received_by = request.user
            purchase.save()
            
            messages.success(request, f'Purchase #{purchase.invoice_no} received successfully! Stock updated.')
            return redirect('core:purchase_list')
    
    return render(request, 'core/purchase_receive.html', {
        'purchase': purchase,
        'items': items,
    })


@login_required
@permission_required('manage_purchases')
def purchase_edit(request, purchase_id):
    """Edit a purchase (only if not received)"""
    purchase = get_object_or_404(Purchase, id=purchase_id)
    suppliers = Supplier.objects.filter(is_active=True)
    reps = Employee.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True)
    
    if purchase.status == 'RECEIVED':
        messages.error(request, 'Cannot edit a received purchase.')
        return redirect('core:purchase_list')
    
    if request.method == 'POST':
        # Update header
        purchase.supplier_id = request.POST.get('supplier')
        purchase.rep_id = request.POST.get('rep') or None
        purchase.po_number = request.POST.get('po_number', '')
        purchase.invoice_no = request.POST.get('invoice_no')
        purchase.purchase_date = request.POST.get('purchase_date') or date.today()
        purchase.due_date = request.POST.get('due_date') or None
        purchase.tax_rate = int(request.POST.get('tax_rate', 0))
        purchase.tax_invoice_no = request.POST.get('tax_invoice_no', '')
        purchase.paid_amount = Decimal(request.POST.get('paid_amount', '0') or '0')
        purchase.payment_status = request.POST.get('payment_status', 'PENDING')
        purchase.notes = request.POST.get('notes', '')
        
        # Delete old items
        purchase.items.all().delete()
        
        # Process new items
        subtotal = Decimal('0')
        item_count = int(request.POST.get('item_count', 0))
        for i in range(item_count):
            product_id = request.POST.get(f'product_{i}')
            quantity = Decimal(request.POST.get(f'qty_{i}', '0') or '0')
            cost_price = Decimal(request.POST.get(f'cost_{i}', '0') or '0')
            retail_price = Decimal(request.POST.get(f'retail_{i}', '0') or '0')
            wholesale_price = Decimal(request.POST.get(f'wholesale_{i}', '0') or '0')
            
            if not product_id or quantity <= 0 or cost_price <= 0:
                continue
            
            product = get_object_or_404(Product, id=product_id)
            total = quantity * cost_price
            retail_total = quantity * retail_price
            wholesale_total = quantity * wholesale_price
            
            PurchaseItem.objects.create(
                purchase=purchase,
                product=product,
                quantity=quantity,
                cost_price=cost_price,
                retail_price=retail_price,
                wholesale_price=wholesale_price,
                total=total,
                retail_total=retail_total,
                wholesale_total=wholesale_total,
            )
            
            subtotal += total
        
        purchase.subtotal = subtotal
        purchase.tax_amount = (subtotal * purchase.tax_rate) / 100
        purchase.total = subtotal + purchase.tax_amount
        purchase.save()
        
        messages.success(request, f'Purchase #{purchase.invoice_no} updated successfully!')
        return redirect('core:purchase_list')
    
    context = {
        'purchase': purchase,
        'items': purchase.items.all().select_related('product'),
        'suppliers': suppliers,
        'reps': reps,
        'products': products,
        'tax_rates': Purchase.TAX_RATE_CHOICES,
    }
    return render(request, 'core/purchase_edit.html', context)


@login_required
@permission_required('manage_purchases')
def purchase_delete(request, purchase_id):
    """Delete a purchase (only if not received)"""
    purchase = get_object_or_404(Purchase, id=purchase_id)
    
    if purchase.status == 'RECEIVED':
        messages.error(request, 'Cannot delete a received purchase.')
        return redirect('core:purchase_list')
    
    if request.method == 'POST':
        purchase.delete()
        messages.success(request, f'Purchase #{purchase.invoice_no} deleted successfully.')
        return redirect('core:purchase_list')
    
    return render(request, 'core/purchase_confirm_delete.html', {'purchase': purchase})


@login_required
@permission_required('manage_purchases')
def supplier_list(request):
    """List all suppliers"""
    suppliers = Supplier.objects.all()
    
    search = request.GET.get('search', '')
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) | 
            Q(contact_person__icontains=search) |
            Q(phone__icontains=search)
        )
    
    context = {
        'suppliers': suppliers,
        'search': search,
        'total_count': suppliers.count(),
    }
    return render(request, 'core/supplier_list.html', context)


@login_required
@permission_required('manage_purchases')
def supplier_add(request):
    """Add a new supplier"""
    if request.method == 'POST':
        name = request.POST.get('name')
        contact_person = request.POST.get('contact_person')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        address = request.POST.get('address')
        tax_number = request.POST.get('tax_number')
        notes = request.POST.get('notes')
        
        if not name:
            messages.error(request, 'Supplier name is required.')
            return render(request, 'core/supplier_form.html', {'action': 'Add', 'supplier': None})
        
        if Supplier.objects.filter(name__iexact=name).exists():
            messages.error(request, f'Supplier "{name}" already exists.')
            return render(request, 'core/supplier_form.html', {'action': 'Add', 'supplier': None})
        
        Supplier.objects.create(
            name=name,
            contact_person=contact_person,
            phone=phone,
            email=email,
            address=address,
            tax_number=tax_number,
            notes=notes,
            is_active=True
        )
        
        messages.success(request, f'Supplier "{name}" added successfully!')
        return redirect('core:supplier_list')
    
    return render(request, 'core/supplier_form.html', {'action': 'Add', 'supplier': None})


@login_required
@permission_required('manage_purchases')
def supplier_edit(request, supplier_id):
    """Edit a supplier"""
    supplier = get_object_or_404(Supplier, id=supplier_id)
    
    if request.method == 'POST':
        supplier.name = request.POST.get('name')
        supplier.contact_person = request.POST.get('contact_person')
        supplier.phone = request.POST.get('phone')
        supplier.email = request.POST.get('email')
        supplier.address = request.POST.get('address')
        supplier.tax_number = request.POST.get('tax_number')
        supplier.notes = request.POST.get('notes')
        supplier.is_active = request.POST.get('is_active') == 'on'
        supplier.save()
        
        messages.success(request, f'Supplier "{supplier.name}" updated successfully!')
        return redirect('core:supplier_list')
    
    return render(request, 'core/supplier_form.html', {'action': 'Edit', 'supplier': supplier})


@login_required
@permission_required('manage_purchases')
def supplier_delete(request, supplier_id):
    """Delete a supplier"""
    supplier = get_object_or_404(Supplier, id=supplier_id)
    
    if supplier.purchases.exists():
        messages.error(request, f'Cannot delete "{supplier.name}" as it has purchase records.')
        return redirect('core:supplier_list')
    
    if request.method == 'POST':
        supplier.delete()
        messages.success(request, f'Supplier "{supplier.name}" deleted successfully.')
        return redirect('core:supplier_list')
    
    return render(request, 'core/supplier_confirm_delete.html', {'supplier': supplier})


@login_required
@permission_required('view_sales')
def cheque_list(request):
    """List all cheques with filters"""
    cheques = Cheque.objects.all().select_related('bank', 'sales_bill')
    
    # Search by Cheque No or Customer
    search = request.GET.get('search', '')
    if search:
        cheques = cheques.filter(
            Q(cheque_no__icontains=search) | 
            Q(customer_name__icontains=search) |
            Q(sales_bill__invoice_no__icontains=search)
        )
    
    # Date filter
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if start_date:
        cheques = cheques.filter(cheque_date__gte=start_date)
    if end_date:
        cheques = cheques.filter(cheque_date__lte=end_date)
    
    # Bank filter
    bank_id = request.GET.get('bank', '')
    if bank_id and bank_id.isdigit():
        cheques = cheques.filter(bank_id=int(bank_id))
    
    # Status filter
    status = request.GET.get('status', '')
    if status:
        cheques = cheques.filter(status=status)
    
    banks = Bank.objects.filter(is_active=True)
    
    # Calculate totals by status
    total_cheques = cheques.count()
    total_amount = cheques.aggregate(total=Sum('amount'))['total'] or 0
    
    pending_amount = cheques.filter(status='PENDING').aggregate(total=Sum('amount'))['total'] or 0
    deposited_amount = cheques.filter(status='DEPOSITED').aggregate(total=Sum('amount'))['total'] or 0
    cleared_amount = cheques.filter(status='CLEARED').aggregate(total=Sum('amount'))['total'] or 0
    bounced_amount = cheques.filter(status='BOUNCED').aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'cheques': cheques,
        'banks': banks,
        'search': search,
        'start_date': start_date,
        'end_date': end_date,
        'selected_bank': bank_id,
        'selected_status': status,
        'total_cheques': total_cheques,
        'total_amount': total_amount,
        'pending_amount': pending_amount,
        'deposited_amount': deposited_amount,
        'cleared_amount': cleared_amount,
        'bounced_amount': bounced_amount,
        'status_choices': Cheque.STATUS_CHOICES,
    }
    return render(request, 'core/cheque_list.html', context)


@login_required
@permission_required('view_sales')
def cheque_detail(request, cheque_id):
    """View cheque details"""
    cheque = get_object_or_404(Cheque, id=cheque_id)
    
    context = {
        'cheque': cheque,
    }
    return render(request, 'core/cheque_detail.html', context)


@login_required
@permission_required('create_sales')
def cheque_deposit(request, cheque_id):
    """Mark cheque as deposited"""
    cheque = get_object_or_404(Cheque, id=cheque_id)
    
    if cheque.status != 'PENDING':
        messages.error(request, f'Cannot deposit a cheque with status: {cheque.get_status_display()}')
        return redirect('core:cheque_list')
    
    if request.method == 'POST':
        deposit_date = request.POST.get('deposit_date')
        notes = request.POST.get('notes', '')
        
        if not deposit_date:
            messages.error(request, 'Please enter a deposit date.')
            return render(request, 'core/cheque_deposit.html', {'cheque': cheque})
        
        cheque.status = 'DEPOSITED'
        cheque.deposit_date = deposit_date
        cheque.notes = notes
        cheque.save()
        
        messages.success(request, f'Cheque #{cheque.cheque_no} marked as DEPOSITED.')
        return redirect('core:cheque_list')
    
    return render(request, 'core/cheque_deposit.html', {'cheque': cheque})


@login_required
@permission_required('create_sales')
def cheque_clear(request, cheque_id):
    """Mark cheque as cleared"""
    cheque = get_object_or_404(Cheque, id=cheque_id)
    
    if cheque.status != 'DEPOSITED':
        messages.error(request, f'Cannot clear a cheque with status: {cheque.get_status_display()}')
        return redirect('core:cheque_list')
    
    if request.method == 'POST':
        cleared_date = request.POST.get('cleared_date')
        notes = request.POST.get('notes', '')
        
        if not cleared_date:
            messages.error(request, 'Please enter a cleared date.')
            return render(request, 'core/cheque_clear.html', {'cheque': cheque})
        
        cheque.status = 'CLEARED'
        cheque.cleared_date = cleared_date
        cheque.notes = notes
        cheque.save()
        
        # Update sales bill payment status if all payments complete
        # This is optional - you can add logic to check if the bill is fully paid
        
        messages.success(request, f'Cheque #{cheque.cheque_no} marked as CLEARED.')
        return redirect('core:cheque_list')
    
    return render(request, 'core/cheque_clear.html', {'cheque': cheque})


@login_required
@permission_required('create_sales')
def cheque_bounce(request, cheque_id):
    cheque = get_object_or_404(Cheque, id=cheque_id)
    bill = cheque.sales_bill

    if request.method == 'POST':
        bounce_reason = request.POST.get('bounce_reason')
        add_bank_charge = request.POST.get('add_bank_charge') == 'on'
        bank_charge_amount = Decimal(request.POST.get('bank_charge_amount', '0'))
        notes = request.POST.get('notes', '')

        if not bounce_reason:
            messages.error(request, 'Please provide a bounce reason.')
            return render(request, 'core/cheque_bounce.html', {'cheque': cheque})

        with transaction.atomic():
            # 1. Mark cheque as bounced
            cheque.status = 'BOUNCED'
            cheque.bounce_reason = bounce_reason
            cheque.bounced_at = timezone.now()
            cheque.bounced_by = request.user
            cheque.notes = notes
            cheque.save()

            # 2. Reverse the payment (find the payment linked to this cheque)
            payment = Payment.objects.filter(
                bill=bill,
                type='Cheque',
                amount=cheque.amount,
                is_reversed=False
            ).first()
            if payment:
                payment.is_reversed = True
                payment.reversed_at = timezone.now()
                payment.reversed_by = request.user
                payment.reversed_cheque = cheque
                payment.save()
            else:
                # If no matching payment, create a reversal record manually? 
                # We'll log a warning and continue.
                messages.warning(request, 'Could not find a matching payment for this cheque.')

            # 3. Restore outstanding balance (by reversing the payment)
            # No need to manually adjust bill – outstanding is computed from payments.

            # 4. Add bank charge expense if requested
            if add_bank_charge and bank_charge_amount > 0:
                # Get or create a default vehicle for bank charges? Use a system vehicle or None.
                # We'll create an expense with no vehicle (optional).
                expense = Expense.objects.create(
                    vehicle=None,
                    date=timezone.now().date(),
                    category='Bank Charges',
                    amount=bank_charge_amount,
                    note=f'Bank charge for bounced cheque {cheque.cheque_no} from {cheque.customer_name}',
                    status='PAID',  # Assume immediately paid
                )
                cheque.bank_charge_amount = bank_charge_amount
                cheque.bank_charge_expense = expense
                cheque.save()
                messages.success(request, f'Bank charge of Rs {bank_charge_amount} added as expense.')

            # 5. Update Credit Collection status (if applicable)
            # If there is a collection record, set it back to PENDING so it can be taken again.
            collection = CreditCollection.objects.filter(sales_bill=bill).first()
            if collection:
                collection.status = 'PENDING'
                collection.date_taken = None
                collection.not_collected_reason = None
                collection.save()

        messages.success(request, f'Cheque {cheque.cheque_no} marked as BOUNCED. Bill {bill.invoice_no} outstanding restored.')
        return redirect('core:cheque_list')

    # GET request: show bounce form
    return render(request, 'core/cheque_bounce.html', {'cheque': cheque})


@login_required
@permission_required('view_reports')
def bounced_cheque_report(request):
    from datetime import date
    start_date = request.GET.get('start_date', date.today().strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', date.today().strftime('%Y-%m-%d'))
    bounced_cheques = Cheque.objects.filter(
        status='BOUNCED',
        bounced_at__date__gte=start_date,
        bounced_at__date__lte=end_date
    ).select_related('sales_bill', 'bank', 'bounced_by')

    total_bounced_amount = bounced_cheques.aggregate(total=Sum('amount'))['total'] or 0
    total_bank_charges = bounced_cheques.exclude(bank_charge_amount=0).aggregate(total=Sum('bank_charge_amount'))['total'] or 0

    context = {
        'cheques': bounced_cheques,
        'start_date': start_date,
        'end_date': end_date,
        'total_bounced_amount': total_bounced_amount,
        'total_bank_charges': total_bank_charges,
        'count': bounced_cheques.count(),
    }
    return render(request, 'core/bounced_cheque_report.html', context)


@login_required
@permission_required('view_sales')
def customer_list(request):
    """List all customers with filters"""
    customers = Customer.objects.all()
    
    # Search
    search = request.GET.get('search', '')
    if search:
        customers = customers.filter(
            Q(name__icontains=search) | 
            Q(code__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Filter by type
    customer_type = request.GET.get('customer_type', '')
    if customer_type:
        customers = customers.filter(customer_type=customer_type)
    
    # Filter by status
    status = request.GET.get('status', '')
    if status == 'active':
        customers = customers.filter(is_active=True)
    elif status == 'inactive':
        customers = customers.filter(is_active=False)
    
    # Filter by credit (has outstanding)
    credit_filter = request.GET.get('credit_filter', '')
    if credit_filter == 'has_credit':
        # Get customers with outstanding balance > 0
        customer_list = []
        for c in customers:
            if c.get_outstanding_balance() > 0:
                customer_list.append(c.id)
        customers = customers.filter(id__in=customer_list)
    elif credit_filter == 'no_credit':
        customer_list = []
        for c in customers:
            if c.get_outstanding_balance() == 0:
                customer_list.append(c.id)
        customers = customers.filter(id__in=customer_list)
    
    # Annotate with calculated fields
    for customer in customers:
        customer.outstanding = customer.get_outstanding_balance()
        customer.total_sales = customer.get_total_sales()
    
    # Statistics
    total_customers = Customer.objects.count()
    active_customers = Customer.objects.filter(is_active=True).count()
    customers_with_credit = 0
    total_outstanding = 0
    
    for c in customers:
        if c.outstanding > 0:
            customers_with_credit += 1
        total_outstanding += c.outstanding
    
    context = {
        'customers': customers,
        'search': search,
        'selected_type': customer_type,
        'selected_status': status,
        'selected_credit_filter': credit_filter,
        'total_customers': total_customers,
        'active_customers': active_customers,
        'customers_with_credit': customers_with_credit,
        'total_outstanding': total_outstanding,
        'customer_types': Customer.CUSTOMER_TYPES,
    }
    return render(request, 'core/customer_list.html', context)


@login_required
@permission_required('manage_products')
def customer_add(request):
    """Add a new customer"""
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code', '')
        customer_type = request.POST.get('customer_type', 'RETAIL')
        phone = request.POST.get('phone', '')
        email = request.POST.get('email', '')
        address = request.POST.get('address', '')
        city = request.POST.get('city', '')
        tax_number = request.POST.get('tax_number', '')
        credit_limit = request.POST.get('credit_limit', '0')
        notes = request.POST.get('notes', '')
        is_active = request.POST.get('is_active') == 'on'
        
        if not name:
            messages.error(request, 'Customer name is required.')
            return render(request, 'core/customer_form.html', {
                'action': 'Add',
                'customer': None,
                'customer_types': Customer.CUSTOMER_TYPES,
            })
        
        # Check if code already exists (if provided)
        if code and Customer.objects.filter(code=code).exists():
            messages.error(request, f'Customer code "{code}" already exists.')
            return render(request, 'core/customer_form.html', {
                'action': 'Add',
                'customer': None,
                'customer_types': Customer.CUSTOMER_TYPES,
            })
        
        # Create customer
        customer = Customer.objects.create(
            name=name,
            code=code if code else '',
            customer_type=customer_type,
            phone=phone,
            email=email,
            address=address,
            city=city,
            tax_number=tax_number,
            credit_limit=credit_limit,
            notes=notes,
            is_active=is_active
        )
        
        messages.success(request, f'Customer "{customer.name}" added successfully! Code: {customer.code}')
        return redirect('core:customer_list')
    
    return render(request, 'core/customer_form.html', {
        'action': 'Add',
        'customer': None,
        'customer_types': Customer.CUSTOMER_TYPES,
    })


@login_required
@permission_required('manage_products')
def customer_edit(request, customer_id):
    """Edit an existing customer"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        customer.name = request.POST.get('name')
        customer.code = request.POST.get('code')
        customer.customer_type = request.POST.get('customer_type')
        customer.phone = request.POST.get('phone', '')
        customer.email = request.POST.get('email', '')
        customer.address = request.POST.get('address', '')
        customer.city = request.POST.get('city', '')
        customer.tax_number = request.POST.get('tax_number', '')
        customer.credit_limit = request.POST.get('credit_limit', '0')
        customer.notes = request.POST.get('notes', '')
        customer.is_active = request.POST.get('is_active') == 'on'
        
        # Check if code is unique (excluding current customer)
        if Customer.objects.filter(code=customer.code).exclude(id=customer.id).exists():
            messages.error(request, f'Customer code "{customer.code}" already exists.')
            return render(request, 'core/customer_form.html', {
                'action': 'Edit',
                'customer': customer,
                'customer_types': Customer.CUSTOMER_TYPES,
            })
        
        customer.save()
        messages.success(request, f'Customer "{customer.name}" updated successfully!')
        return redirect('core:customer_list')
    
    return render(request, 'core/customer_form.html', {
        'action': 'Edit',
        'customer': customer,
        'customer_types': Customer.CUSTOMER_TYPES,
    })


@login_required
@permission_required('manage_products')
def customer_delete(request, customer_id):
    """Delete a customer (with confirmation)"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    # Check if customer has sales
    if SalesBill.objects.filter(shop_code=customer.code).exists():
        messages.error(request, f'Cannot delete "{customer.name}" as it has sales records.')
        return redirect('core:customer_list')
    
    if request.method == 'POST':
        customer.delete()
        messages.success(request, f'Customer "{customer.name}" deleted successfully.')
        return redirect('core:customer_list')
    
    return render(request, 'core/customer_confirm_delete.html', {'customer': customer})


@login_required
@permission_required('manage_products')
def customer_toggle_status(request, customer_id):
    """Activate or deactivate a customer"""
    customer = get_object_or_404(Customer, id=customer_id)
    customer.is_active = not customer.is_active
    customer.save()
    status = 'activated' if customer.is_active else 'deactivated'
    messages.success(request, f'Customer "{customer.name}" has been {status}.')
    return redirect('core:customer_list')


@login_required
@permission_required('view_sales')
def customer_detail(request, customer_id):
    """View customer details with sales history"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    # Get sales bills for this customer
    sales_bills = SalesBill.objects.filter(shop_code=customer.code).order_by('-date', '-created_at')
    
    # Get payment history
    payments = Payment.objects.filter(bill__shop_code=customer.code).select_related('bill')
    
    # Calculate totals
    total_sales = customer.get_total_sales()
    outstanding = customer.get_outstanding_balance()
    
    context = {
        'customer': customer,
        'sales_bills': sales_bills,
        'payments': payments,
        'total_sales': total_sales,
        'outstanding': outstanding,
        'available_credit': customer.get_available_credit(),
    }
    return render(request, 'core/customer_detail.html', context)


@login_required
@permission_required('manage_customers')
def customer_search_api(request):
    query = request.GET.get('q', '')
    customer_type = request.GET.get('type', '')
    status = request.GET.get('status', '')
    
    customers = Customer.objects.all()
    
    if query and len(query) >= 2:
        customers = customers.filter(
            Q(name__icontains=query) | 
            Q(code__icontains=query) |
            Q(phone__icontains=query)
        )
    
    if customer_type:
        customers = customers.filter(customer_type=customer_type)
    
    if status == 'active':
        customers = customers.filter(is_active=True)
    elif status == 'inactive':
        customers = customers.filter(is_active=False)
    
    customers = customers[:50]
    
    data = []
    for c in customers:
        data.append({
            'id': c.id,
            'code': c.code,
            'name': c.name,
            'customer_type': c.get_customer_type_display(),
            'phone': c.phone or '',
            'address': c.address or '',
            'city': c.city or '',
            'is_active': c.is_active,
            'total_sales': float(c.get_total_sales()),
            'outstanding': float(c.get_outstanding_balance()),
        })
    return JsonResponse({'customers': data})


@login_required
@permission_required('manage_customers')
def customer_check_duplicate(request):
    """Check if a customer with the same name or code already exists."""
    name = request.GET.get('name', '').strip()
    code = request.GET.get('code', '').strip()
    exclude_id = request.GET.get('exclude_id', '')
    
    if not name and not code:
        return JsonResponse({'exists': False})
    
    q = Q()
    if name:
        q |= Q(name__iexact=name)
    if code:
        q |= Q(code__iexact=code)
    
    customers = Customer.objects.filter(q)
    if exclude_id and exclude_id.isdigit():
        customers = customers.exclude(id=int(exclude_id))
    
    exists = customers.exists()
    duplicate = customers.first()
    
    return JsonResponse({
        'exists': exists,
        'duplicate': {
            'id': duplicate.id if duplicate else None,
            'name': duplicate.name if duplicate else None,
            'code': duplicate.code if duplicate else None,
        } if duplicate else None,
    })


@login_required
@permission_required('create_sales')
def create_sales_bill(request):
    MAX_ITEMS = 10
    
    if request.method == 'POST':
        form = SalesBillForm(request.POST)
        
        if form.is_valid():
            with transaction.atomic():
                bill = form.save(commit=False)
                
                subtotal = Decimal('0')
                items_to_save = []
                vehicle = bill.vehicle
                
                # Get customer from form (either existing customer ID or new customer name)
                customer_id = request.POST.get('customer')
                customer_name = request.POST.get('shop_name', '').strip()
                customer_code = request.POST.get('shop_code', '').strip()
                
                if customer_id and customer_id.isdigit():
                    try:
                        customer = Customer.objects.get(id=customer_id)
                        bill.shop_name = customer.name
                        bill.shop_code = customer.code
                    except Customer.DoesNotExist:
                        if customer_name:
                            bill.shop_name = customer_name
                            bill.shop_code = customer_code
                else:
                    if customer_name:
                        bill.shop_name = customer_name
                    if customer_code:
                        bill.shop_code = customer_code
                    else:
                        bill.shop_code = ''
                
                # Process items from the POST data (with discount support)
                for i in range(1, MAX_ITEMS + 1):
                    product_id = request.POST.get(f'product_{i}')
                    quantity_str = request.POST.get(f'qty_{i}', '0')
                    rate_str = request.POST.get(f'rate_{i}', '0')
                    is_foc_str = request.POST.get(f'is_foc_{i}', 'false')
                    is_foc = is_foc_str.lower() == 'true'
                    
                    # Discount fields
                    discount_type = request.POST.get(f'discount_type_{i}', '')
                    discount_value_str = request.POST.get(f'discount_value_{i}', '0')
                    
                    if not product_id or not quantity_str or quantity_str == '0':
                        continue
                    
                    product = get_object_or_404(Product, id=product_id)
                    quantity = Decimal(quantity_str)
                    rate = Decimal(rate_str)
                    discount_value = Decimal(discount_value_str) if discount_value_str else Decimal('0')
                    
                    # Calculate discounted rate
                    discounted_rate = rate
                    if discount_type == 'PERCENTAGE' and discount_value > 0:
                        discounted_rate = rate - (rate * discount_value / 100)
                    elif discount_type == 'FIXED' and discount_value > 0:
                        discounted_rate = rate - discount_value
                    
                    if discounted_rate < 0:
                        discounted_rate = Decimal('0')
                    
                    total = quantity * discounted_rate
                    
                    # Check Vehicle Stock (for sales, not returns)
                    if quantity > 0:
                        vehicle_stock = get_object_or_404(VehicleStock, vehicle=vehicle, product=product)
                        if vehicle_stock.quantity < quantity:
                            messages.error(request, f'❌ Not enough stock on {vehicle.vehicle_number} for {product.name}. Available: {vehicle_stock.quantity}')
                            customers = Customer.objects.filter(is_active=True)
                            return render(request, 'core/sales_bill.html', {
                                'form': form,
                                'products': Product.objects.all(),
                                'customers': customers,
                                'vehicles': Vehicle.objects.filter(is_active=True),
                                'reps': Employee.objects.filter(position='Rep', is_active=True),
                                'banks': Bank.objects.filter(is_active=True),
                                'max_items': MAX_ITEMS,
                            })
                        # Deduct from Vehicle Stock
                        vehicle_stock.quantity -= quantity
                        vehicle_stock.save()
                    elif quantity < 0:
                        # RETURN: Add back to Vehicle Stock
                        vehicle_stock, created = VehicleStock.objects.get_or_create(
                            vehicle=vehicle, 
                            product=product
                        )
                        vehicle_stock.quantity += abs(quantity)
                        vehicle_stock.save()
                    
                    subtotal += total
                    items_to_save.append({
                        'product': product,
                        'quantity': quantity,
                        'rate': rate,
                        'discounted_rate': discounted_rate,
                        'total': total,
                        'is_foc': is_foc,
                        'discount_type': discount_type if discount_value > 0 else None,
                        'discount_value': discount_value if discount_value > 0 else Decimal('0'),
                    })
                
                # Calculate totals (subtotal already calculated)
                bill.subtotal = subtotal
                bill.discount_total = Decimal('0')  # We'll use bill_discount_amount instead
                
                # Process bill-level discount
                bill_discount_type = request.POST.get('bill_discount_type', '')
                bill_discount_value_str = request.POST.get('bill_discount_value', '0')
                bill_discount_value = Decimal(bill_discount_value_str) if bill_discount_value_str else Decimal('0')
                
                if bill_discount_type == 'PERCENTAGE' and bill_discount_value > 0:
                    bill_discount_amount = (subtotal * bill_discount_value) / 100
                elif bill_discount_type == 'FIXED' and bill_discount_value > 0:
                    bill_discount_amount = bill_discount_value
                    if bill_discount_amount > subtotal:
                        bill_discount_amount = subtotal  # Can't discount more than subtotal
                else:
                    bill_discount_amount = Decimal('0')
                
                bill.bill_discount_type = bill_discount_type if bill_discount_value > 0 else None
                bill.bill_discount_value = bill_discount_value if bill_discount_value > 0 else Decimal('0')
                bill.bill_discount_amount = bill_discount_amount
                bill.net_total = subtotal - bill_discount_amount
                
                # Set bill status from form (COMPLETED or DRAFT)
                bill.status = request.POST.get('bill_status', 'DRAFT')
                bill.is_return = request.POST.get('is_return') == 'true'
                bill.return_reason = request.POST.get('return_reason') or None
                
                # Set rep if provided
                rep_id = request.POST.get('rep')
                if rep_id and rep_id.isdigit():
                    try:
                        bill.rep = Employee.objects.get(id=rep_id, position='Rep')
                    except Employee.DoesNotExist:
                        pass
                
                # Invoice number handling
                user_invoice = request.POST.get('invoice_no', '').strip()
                if not user_invoice:
                    messages.error(request, '❌ Invoice Number is required. Please enter a unique invoice number.')
                    customers = Customer.objects.filter(is_active=True)
                    return render(request, 'core/sales_bill.html', {
                        'form': form,
                        'products': Product.objects.all(),
                        'customers': customers,
                        'vehicles': Vehicle.objects.filter(is_active=True),
                        'reps': Employee.objects.filter(position='Rep', is_active=True),
                        'banks': Bank.objects.filter(is_active=True),
                        'max_items': MAX_ITEMS,
                    })
                
                # Check duplicate invoice number
                if SalesBill.objects.filter(invoice_no=user_invoice).exists():
                    messages.error(request, f'❌ Invoice number "{user_invoice}" already exists. Please use a unique number.')
                    customers = Customer.objects.filter(is_active=True)
                    return render(request, 'core/sales_bill.html', {
                        'form': form,
                        'products': Product.objects.all(),
                        'customers': customers,
                        'vehicles': Vehicle.objects.filter(is_active=True),
                        'reps': Employee.objects.filter(position='Rep', is_active=True),
                        'banks': Bank.objects.filter(is_active=True),
                        'max_items': MAX_ITEMS,
                        'selected_vehicle': request.POST.get('vehicle', ''),
                        'selected_rep': request.POST.get('rep', ''),
                        'shop_name': request.POST.get('shop_name', ''),
                        'shop_code': request.POST.get('shop_code', ''),
                        'customer_id': request.POST.get('customer', ''),
                        'random_invoice': user_invoice,
                        'duplicate_error': True,
                    })
                
                bill.invoice_no = user_invoice
                bill.save()
                
                # If status is COMPLETED, set completed_by and completed_at
                if bill.status == 'COMPLETED':
                    bill.completed_by = request.user
                    bill.completed_at = timezone.now()
                    bill.save()
                
                # 4. Save Item lines (with discount info)
                for item_data in items_to_save:
                    SalesItem.objects.create(
                        bill=bill,
                        product=item_data['product'],
                        quantity=item_data['quantity'],
                        rate=item_data['rate'],
                        discounted_rate=item_data['discounted_rate'],
                        total=item_data['total'],
                        is_foc=item_data['is_foc'],
                        discount_type=item_data['discount_type'],
                        discount_value=item_data['discount_value'],
                    )
                
                # 5. Process Payments (Cash, Credit, Cheque, Online, Multi)
                cash = Decimal(request.POST.get('cash_amount', '0') or '0')
                credit = Decimal(request.POST.get('credit_amount', '0') or '0')
                cheque = Decimal(request.POST.get('cheque_amount', '0') or '0')
                online = Decimal(request.POST.get('online_amount', '0') or '0')
                total_paid = cash + credit + cheque + online
                
                if total_paid != bill.net_total:
                    messages.error(request, f'❌ Payment total ({total_paid}) does not match Bill Total ({bill.net_total})!')
                    customers = Customer.objects.filter(is_active=True)
                    return render(request, 'core/sales_bill.html', {
                        'form': form,
                        'products': Product.objects.all(),
                        'customers': customers,
                        'vehicles': Vehicle.objects.filter(is_active=True),
                        'reps': Employee.objects.filter(position='Rep', is_active=True),
                        'banks': Bank.objects.filter(is_active=True),
                        'max_items': MAX_ITEMS,
                    })
                
                if cash > 0:
                    Payment.objects.create(bill=bill, type='Cash', amount=cash)
                
                if credit > 0:
                    Payment.objects.create(bill=bill, type='Credit', amount=credit)
                
                cheque_amount = Decimal(request.POST.get('cheque_amount', '0') or '0')
                if cheque_amount > 0:
                    Payment.objects.create(
                        bill=bill,
                        type='Cheque',
                        amount=cheque_amount
                    )
                    cheque_no = request.POST.get('cheque_no', '')
                    cheque_date = request.POST.get('cheque_date', '')
                    bank_id = request.POST.get('cheque_bank', '')
                    if cheque_no and cheque_date and bank_id:
                        try:
                            bank = Bank.objects.get(id=bank_id)
                            Cheque.objects.create(
                                cheque_no=cheque_no,
                                bank=bank,
                                cheque_date=cheque_date,
                                amount=cheque_amount,
                                customer_name=bill.shop_name or bill.shop_code or 'N/A',
                                sales_bill=bill,
                                status='PENDING',
                                notes=f"Auto-created from invoice: {bill.invoice_no}"
                            )
                        except Bank.DoesNotExist:
                            pass
                
                if online > 0:
                    Payment.objects.create(bill=bill, type='Online', amount=online)
                    online_method = request.POST.get('online_method', '')
                    online_ref = request.POST.get('online_ref', '')
                    if online_method:
                        OnlinePayment.objects.create(
                            bill=bill,
                            payment_method=online_method,
                            reference_no=online_ref,
                            amount=online
                        )
                
                # Multi Pay
                multi_type1 = request.POST.get('multi_type1', '')
                multi_amount1 = Decimal(request.POST.get('multi_amount1', '0') or '0')
                multi_type2 = request.POST.get('multi_type2', '')
                multi_amount2 = Decimal(request.POST.get('multi_amount2', '0') or '0')
                if multi_type1 and multi_type2 and multi_amount1 > 0 and multi_amount2 > 0:
                    Payment.objects.create(
                        bill=bill,
                        type=multi_type1,
                        amount=multi_amount1
                    )
                    Payment.objects.create(
                        bill=bill,
                        type=multi_type2,
                        amount=multi_amount2
                    )
                                
                # ============================================================
                # ✅ SESSION SAVING – AFTER ALL SAVES
                # ============================================================
                try:
                    if bill.vehicle:
                        request.session['last_vehicle_id'] = int(bill.vehicle.id)
                    else:
                        request.session['last_vehicle_id'] = None
                    
                    if bill.rep:
                        request.session['last_rep_id'] = int(bill.rep.id)
                    else:
                        request.session['last_rep_id'] = None
                    
                    request.session.modified = True
                except Exception as e:
                    pass
                
                # ✅ Success message and redirect
                messages.success(request, f'✅ Bill {bill.invoice_no} saved successfully! Total: {bill.net_total}')
                return redirect('/sales/')
        
        else:
            # Form is invalid – re-render with errors
            messages.error(request, 'Please correct the errors below.')
            customers = Customer.objects.filter(is_active=True)
            products = Product.objects.all()
            vehicles = Vehicle.objects.filter(is_active=True)
            reps = Employee.objects.filter(position='Rep', is_active=True)
            banks = Bank.objects.filter(is_active=True)
            context = {
                'form': form,
                'products': products,
                'customers': customers,
                'vehicles': vehicles,
                'reps': reps,
                'banks': banks,
                'max_items': MAX_ITEMS,
                'today': date.today(),
                'random_invoice': request.POST.get('invoice_no', ''),
                'selected_vehicle': request.POST.get('vehicle', ''),
                'selected_rep': request.POST.get('rep', ''),
                'shop_name': request.POST.get('shop_name', ''),
                'shop_code': request.POST.get('shop_code', ''),
                'customer_id': request.POST.get('customer', ''),
                'cart_items': [],
                'return_mode': False,
                'duplicate_error': False,
            }
            return render(request, 'core/sales_bill.html', context)
    
    else:
        form = SalesBillForm()
    
    # GET request: Show empty form
    products = Product.objects.filter(is_active=True)
    customers = Customer.objects.filter(is_active=True)
    vehicles = Vehicle.objects.filter(is_active=True)
    reps = Employee.objects.filter(position='Rep', is_active=True)
    banks = Bank.objects.filter(is_active=True)
    
    # Try to load from session
    selected_vehicle = request.GET.get('vehicle')
    selected_rep = request.GET.get('rep')

    # If no GET parameter, load from session
    if not selected_vehicle:
        selected_vehicle = request.session.get('last_vehicle_id')
    if not selected_rep:
        selected_rep = request.session.get('last_rep_id')

    # ... later in the GET section ...

    # Get vehicle stock for the selected vehicle
    vehicle_stock_dict = {}
    if selected_vehicle:
        try:
            # Convert to int safely (works for both string and int values)
            vehicle_id = int(selected_vehicle)
            vehicle = Vehicle.objects.get(id=vehicle_id)
            vehicle_stocks = VehicleStock.objects.filter(vehicle=vehicle)
            for stock in vehicle_stocks:
                vehicle_stock_dict[stock.product_id] = stock.quantity
        except (ValueError, TypeError, Vehicle.DoesNotExist):
            # If conversion fails or vehicle not found, ignore
            pass
    
    product_list = []
    for product in products:
        product_list.append({
            'id': product.id,
            'name': product.name,
            'code': product.code,
            'selling_price': product.selling_price,
            'unit': product.unit,
            'vehicle_stock': vehicle_stock_dict.get(product.id, 0),
        })
    
    random_invoice = request.GET.get('invoice_no', '')
    context = {
        'form': form,
        'products': product_list,  # Use product_list with stock info
        'customers': customers,
        'vehicles': vehicles,
        'reps': reps,
        'banks': banks,
        'max_items': MAX_ITEMS,
        'today': date.today(),
        'random_invoice': random_invoice,
        'selected_vehicle': selected_vehicle,
        'selected_rep': selected_rep,
        'shop_name': request.GET.get('shop_name', ''),
        'shop_code': request.GET.get('shop_code', ''),
        'customer_id': request.GET.get('customer_id', ''),
        'cart_items': [],
        'return_mode': False,
        'duplicate_error': False,
    }
    return render(request, 'core/sales_bill.html', context)


@login_required
@permission_required('view_expenses')
def expense_list(request):
    """List all expenses with filters"""
    expenses = Expense.objects.all().select_related('vehicle', 'employee', 'approved_by')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        expenses = expenses.filter(
            Q(category__icontains=search) |
            Q(note__icontains=search) |
            Q(invoice_no__icontains=search)
        )
    
    # Date filter
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if start_date:
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        expenses = expenses.filter(date__lte=end_date)
    
    # Category filter
    category = request.GET.get('category', '')
    if category:
        expenses = expenses.filter(category=category)
    
    # Vehicle filter
    vehicle_id = request.GET.get('vehicle', '')
    if vehicle_id and vehicle_id.isdigit():
        expenses = expenses.filter(vehicle_id=int(vehicle_id))
    
    # Employee filter
    employee_id = request.GET.get('employee', '')
    if employee_id and employee_id.isdigit():
        expenses = expenses.filter(employee_id=int(employee_id))
    
    # Status filter
    status = request.GET.get('status', '')
    if status:
        expenses = expenses.filter(status=status)
    
    # Payment method filter
    payment_method = request.GET.get('payment_method', '')
    if payment_method:
        expenses = expenses.filter(payment_method=payment_method)
    
    # ✅ Calculate totals with safe aggregation
    total_expenses = 0
    for exp in expenses:
        try:
            total_expenses += float(exp.amount)
        except:
            pass
    
    # Category-wise breakdown
    category_breakdown = []
    for exp in expenses:
        try:
            amount = float(exp.amount)
        except:
            amount = 0
        # Simple manual grouping
        found = False
        for item in category_breakdown:
            if item['category'] == exp.category:
                item['total'] += amount
                found = True
                break
        if not found:
            category_breakdown.append({
                'category': exp.category,
                'total': amount
            })
    category_breakdown.sort(key=lambda x: x['total'], reverse=True)
    
    # Status-wise breakdown
    status_breakdown = []
    for exp in expenses:
        try:
            amount = float(exp.amount)
        except:
            amount = 0
        found = False
        for item in status_breakdown:
            if item['status'] == exp.status:
                item['total'] += amount
                item['count'] += 1
                found = True
                break
        if not found:
            status_breakdown.append({
                'status': exp.status,
                'total': amount,
                'count': 1
            })
    
    # Vehicles list for filter
    vehicles = Vehicle.objects.filter(is_active=True)
    employees = Employee.objects.filter(is_active=True)
    
    context = {
        'expenses': expenses,
        'search': search,
        'start_date': start_date,
        'end_date': end_date,
        'selected_category': category,
        'selected_vehicle': vehicle_id,
        'selected_employee': employee_id,
        'selected_status': status,
        'selected_payment_method': payment_method,
        'total_expenses': total_expenses,
        'category_breakdown': category_breakdown,
        'status_breakdown': status_breakdown,
        'vehicles': vehicles,
        'employees': employees,
        'categories': Expense.CATEGORY_CHOICES,
        'payment_methods': Expense.PAYMENT_METHODS,
        'status_choices': Expense.STATUS_CHOICES,
    }
    return render(request, 'core/expense_list.html', context)


@login_required
@permission_required('manage_expenses')
def expense_add(request):
    """Add a new expense"""
    vehicles = Vehicle.objects.filter(is_active=True)
    employees = Employee.objects.filter(is_active=True)
    
    if request.method == 'POST':
        # Get form data
        expense_date = request.POST.get('date')  # Renamed to expense_date
        category = request.POST.get('category')
        vehicle_id = request.POST.get('vehicle')
        employee_id = request.POST.get('employee')
        amount_str = request.POST.get('amount')
        payment_method = request.POST.get('payment_method', 'Cash')
        invoice_no = request.POST.get('invoice_no', '')
        note = request.POST.get('note', '')
        status = request.POST.get('status', 'PENDING')
        
        # Validation
        errors = []
        if not expense_date:
            errors.append('Please select a date.')
        if not category:
            errors.append('Please select a category.')
        if not amount_str:
            errors.append('Please enter an amount.')
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'core/expense_form.html', {
                'action': 'Add',
                'expense': None,
                'vehicles': vehicles,
                'employees': employees,
                'categories': Expense.CATEGORY_CHOICES,
                'payment_methods': Expense.PAYMENT_METHODS,
                'status_choices': Expense.STATUS_CHOICES,
                'today': date.today(),
            })
        
        # Validate amount
        try:
            amount = Decimal(amount_str)
            if amount <= 0:
                messages.error(request, 'Amount must be greater than 0.')
                return render(request, 'core/expense_form.html', {
                    'action': 'Add',
                    'expense': None,
                    'vehicles': vehicles,
                    'employees': employees,
                    'categories': Expense.CATEGORY_CHOICES,
                    'payment_methods': Expense.PAYMENT_METHODS,
                    'status_choices': Expense.STATUS_CHOICES,
                    'today': date.today(),
                })
        except:
            messages.error(request, 'Please enter a valid amount.')
            return render(request, 'core/expense_form.html', {
                'action': 'Add',
                'expense': None,
                'vehicles': vehicles,
                'employees': employees,
                'categories': Expense.CATEGORY_CHOICES,
                'payment_methods': Expense.PAYMENT_METHODS,
                'status_choices': Expense.STATUS_CHOICES,
                'today': date.today(),
            })
        
        # Get vehicle and employee
        vehicle = None
        if vehicle_id and vehicle_id.isdigit():
            vehicle = get_object_or_404(Vehicle, id=vehicle_id)
        
        employee = None
        if employee_id and employee_id.isdigit():
            employee = get_object_or_404(Employee, id=employee_id)
        
        # Create expense
        expense = Expense.objects.create(
            date=expense_date,
            category=category,
            vehicle=vehicle,
            employee=employee,
            amount=amount,
            payment_method=payment_method,
            invoice_no=invoice_no,
            note=note,
            status=status,
        )
        
        messages.success(request, f'Expense added successfully!')
        return redirect('core:expense_list')
    
    # GET request
    return render(request, 'core/expense_form.html', {
        'action': 'Add',
        'expense': None,
        'vehicles': vehicles,
        'employees': employees,
        'categories': Expense.CATEGORY_CHOICES,
        'payment_methods': Expense.PAYMENT_METHODS,
        'status_choices': Expense.STATUS_CHOICES,
        'today': date.today(),
    })


@login_required
@permission_required('manage_expenses')
def expense_edit(request, expense_id):
    """Edit an existing expense"""
    expense = get_object_or_404(Expense, id=expense_id)
    vehicles = Vehicle.objects.filter(is_active=True)
    employees = Employee.objects.filter(is_active=True)
    
    if request.method == 'POST':
        expense_date = request.POST.get('date')
        category = request.POST.get('category')
        vehicle_id = request.POST.get('vehicle')
        employee_id = request.POST.get('employee')
        amount_str = request.POST.get('amount')
        payment_method = request.POST.get('payment_method', 'Cash')
        invoice_no = request.POST.get('invoice_no', '')
        note = request.POST.get('note', '')
        status = request.POST.get('status', 'PENDING')
        
        # Validation
        errors = []
        if not expense_date:
            errors.append('Please select a date.')
        if not category:
            errors.append('Please select a category.')
        if not amount_str:
            errors.append('Please enter an amount.')
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'core/expense_form.html', {
                'action': 'Edit',
                'expense': expense,
                'vehicles': vehicles,
                'employees': employees,
                'categories': Expense.CATEGORY_CHOICES,
                'payment_methods': Expense.PAYMENT_METHODS,
                'status_choices': Expense.STATUS_CHOICES,
                'today': date.today(),
            })
        
        # Validate amount
        try:
            amount = Decimal(amount_str)
            if amount <= 0:
                messages.error(request, 'Amount must be greater than 0.')
                return render(request, 'core/expense_form.html', {
                    'action': 'Edit',
                    'expense': expense,
                    'vehicles': vehicles,
                    'employees': employees,
                    'categories': Expense.CATEGORY_CHOICES,
                    'payment_methods': Expense.PAYMENT_METHODS,
                    'status_choices': Expense.STATUS_CHOICES,
                    'today': date.today(),
                })
        except:
            messages.error(request, 'Please enter a valid amount.')
            return render(request, 'core/expense_form.html', {
                'action': 'Edit',
                'expense': expense,
                'vehicles': vehicles,
                'employees': employees,
                'categories': Expense.CATEGORY_CHOICES,
                'payment_methods': Expense.PAYMENT_METHODS,
                'status_choices': Expense.STATUS_CHOICES,
                'today': date.today(),
            })
        
        # Get vehicle and employee
        vehicle = None
        if vehicle_id and vehicle_id.isdigit():
            vehicle = get_object_or_404(Vehicle, id=vehicle_id)
        
        employee = None
        if employee_id and employee_id.isdigit():
            employee = get_object_or_404(Employee, id=employee_id)
        
        # Update expense
        expense.date = expense_date
        expense.category = category
        expense.vehicle = vehicle
        expense.employee = employee
        expense.amount = amount
        expense.payment_method = payment_method
        expense.invoice_no = invoice_no
        expense.note = note
        expense.status = status
        expense.save()
        
        messages.success(request, f'Expense updated successfully!')
        return redirect('core:expense_list')
    
    # GET request
    return render(request, 'core/expense_form.html', {
        'action': 'Edit',
        'expense': expense,
        'vehicles': vehicles,
        'employees': employees,
        'categories': Expense.CATEGORY_CHOICES,
        'payment_methods': Expense.PAYMENT_METHODS,
        'status_choices': Expense.STATUS_CHOICES,
        'today': date.today(),
    })


@login_required
@permission_required('manage_expenses')
def expense_detail(request, expense_id):
    """View expense details"""
    expense = get_object_or_404(Expense, id=expense_id)
    return render(request, 'core/expense_detail.html', {'expense': expense})


@login_required
@permission_required('manage_expenses')
def expense_delete(request, expense_id):
    """Delete an expense (with confirmation)"""
    expense = get_object_or_404(Expense, id=expense_id)
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Expense deleted successfully.')
        return redirect('core:expense_list')
    return render(request, 'core/expense_confirm_delete.html', {'expense': expense})


@login_required
@permission_required('manage_expenses')
def expense_approve(request, expense_id):
    """Approve an expense"""
    expense = get_object_or_404(Expense, id=expense_id)
    if expense.status != 'PENDING':
        messages.error(request, 'Only pending expenses can be approved.')
        return redirect('core:expense_list')
    
    expense.status = 'APPROVED'
    expense.approved_by = request.user
    expense.approved_at = timezone.now()
    expense.save()
    messages.success(request, f'Expense #{expense.id} approved successfully!')
    return redirect('core:expense_list')


@login_required
@permission_required('manage_expenses')
def expense_pay(request, expense_id):
    """Mark expense as paid"""
    expense = get_object_or_404(Expense, id=expense_id)
    if expense.status not in ['APPROVED', 'PENDING']:
        messages.error(request, 'Only approved or pending expenses can be marked as paid.')
        return redirect('core:expense_list')
    
    expense.status = 'PAID'
    expense.paid_at = timezone.now()
    expense.save()
    messages.success(request, f'Expense #{expense.id} marked as paid!')
    return redirect('core:expense_list')


@login_required
@permission_required('manage_expenses')
def expense_reject(request, expense_id):
    """Reject an expense with reason"""
    expense = get_object_or_404(Expense, id=expense_id)
    if expense.status != 'PENDING':
        messages.error(request, 'Only pending expenses can be rejected.')
        return redirect('core:expense_list')
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        if not reason:
            messages.error(request, 'Please provide a rejection reason.')
            return render(request, 'core/expense_reject.html', {'expense': expense})
        
        expense.status = 'REJECTED'
        expense.rejected_reason = reason
        expense.save()
        messages.warning(request, f'Expense #{expense.id} rejected.')
        return redirect('core:expense_list')
    
    return render(request, 'core/expense_reject.html', {'expense': expense})


@login_required
def expense_summary_api(request):
    """API endpoint for dashboard expense summary"""
    today = date.today()
    start_of_month = date(today.year, today.month, 1)
    
    today_expenses = Expense.objects.filter(date=today).aggregate(total=Sum('amount'))['total'] or 0
    month_expenses = Expense.objects.filter(date__gte=start_of_month, date__lte=today).aggregate(total=Sum('amount'))['total'] or 0
    
    return JsonResponse({
        'today_expense': today_expenses,
        'month_expense': month_expenses,
    })


@login_required
@permission_required('view_vehicles')
def vehicle_list(request):
    """List all vehicles with filters"""
    vehicles = Vehicle.objects.all()
    
    # Search
    search = request.GET.get('search', '')
    if search:
        vehicles = vehicles.filter(
            Q(vehicle_number__icontains=search) |
            Q(driver_name__icontains=search) |
            Q(registration_number__icontains=search)
        )
    
    # Filter by type
    vehicle_type = request.GET.get('vehicle_type', '')
    if vehicle_type:
        vehicles = vehicles.filter(vehicle_type=vehicle_type)
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        vehicles = vehicles.filter(status=status)
    
    # Filter by fuel type
    fuel_type = request.GET.get('fuel_type', '')
    if fuel_type:
        vehicles = vehicles.filter(fuel_type=fuel_type)
    
    # Statistics
    total_vehicles = vehicles.count()
    active_vehicles = vehicles.filter(status='ACTIVE').count()
    maintenance_vehicles = vehicles.filter(status='MAINTENANCE').count()
    
    # Expiry alerts
    from datetime import date, timedelta
    today = date.today()
    insurance_expiring = vehicles.filter(
        insurance_expiry__gte=today,
        insurance_expiry__lte=today + timedelta(days=30)
    ).count()
    registration_expiring = vehicles.filter(
        registration_expiry__gte=today,
        registration_expiry__lte=today + timedelta(days=30)
    ).count()
    
    context = {
        'vehicles': vehicles,
        'search': search,
        'selected_type': vehicle_type,
        'selected_status': status,
        'selected_fuel': fuel_type,
        'total_vehicles': total_vehicles,
        'active_vehicles': active_vehicles,
        'maintenance_vehicles': maintenance_vehicles,
        'insurance_expiring': insurance_expiring,
        'registration_expiring': registration_expiring,
        'vehicle_types': Vehicle.VEHICLE_TYPES,
        'fuel_types': Vehicle.FUEL_TYPES,
        'status_choices': Vehicle.STATUS_CHOICES,
    }
    return render(request, 'core/vehicle_list.html', context)


@login_required
@permission_required('manage_vehicles')
def vehicle_add(request):
    """Add a new vehicle"""
    if request.method == 'POST':
        vehicle_number = request.POST.get('vehicle_number')
        vehicle_type = request.POST.get('vehicle_type')
        driver_name = request.POST.get('driver_name')
        status = request.POST.get('status', 'ACTIVE')
        registration_number = request.POST.get('registration_number', '')
        registration_expiry = request.POST.get('registration_expiry') or None
        insurance_expiry = request.POST.get('insurance_expiry') or None
        last_service_date = request.POST.get('last_service_date') or None
        mileage = request.POST.get('mileage') or None
        fuel_type = request.POST.get('fuel_type', 'Diesel')
        capacity = request.POST.get('capacity') or None
        notes = request.POST.get('notes', '')
        purchase_date = request.POST.get('purchase_date') or None
        purchase_price = request.POST.get('purchase_price') or None
        current_value = request.POST.get('current_value') or None
        is_active = request.POST.get('is_active') == 'on'
        
        if not vehicle_number or not driver_name:
            messages.error(request, 'Vehicle Number and Driver Name are required.')
            return render(request, 'core/vehicle_form.html', {
                'action': 'Add',
                'vehicle': None,
                'vehicle_types': Vehicle.VEHICLE_TYPES,
                'fuel_types': Vehicle.FUEL_TYPES,
                'status_choices': Vehicle.STATUS_CHOICES,
            })
        
        if Vehicle.objects.filter(vehicle_number=vehicle_number).exists():
            messages.error(request, f'Vehicle "{vehicle_number}" already exists.')
            return render(request, 'core/vehicle_form.html', {
                'action': 'Add',
                'vehicle': None,
                'vehicle_types': Vehicle.VEHICLE_TYPES,
                'fuel_types': Vehicle.FUEL_TYPES,
                'status_choices': Vehicle.STATUS_CHOICES,
            })
        
        vehicle = Vehicle.objects.create(
            vehicle_number=vehicle_number,
            vehicle_type=vehicle_type,
            driver_name=driver_name,
            status=status,
            registration_number=registration_number,
            registration_expiry=registration_expiry,
            insurance_expiry=insurance_expiry,
            last_service_date=last_service_date,
            mileage=mileage,
            fuel_type=fuel_type,
            capacity=capacity,
            notes=notes,
            purchase_date=purchase_date,
            purchase_price=purchase_price,
            current_value=current_value,
            is_active=is_active,
        )
        
        messages.success(request, f'Vehicle "{vehicle.vehicle_number}" added successfully!')
        return redirect('core:vehicle_list')
    
    return render(request, 'core/vehicle_form.html', {
        'action': 'Add',
        'vehicle': None,
        'vehicle_types': Vehicle.VEHICLE_TYPES,
        'fuel_types': Vehicle.FUEL_TYPES,
        'status_choices': Vehicle.STATUS_CHOICES,
    })


@login_required
@permission_required('manage_vehicles')
def vehicle_edit(request, vehicle_id):
    """Edit an existing vehicle"""
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    
    if request.method == 'POST':
        vehicle.vehicle_number = request.POST.get('vehicle_number')
        vehicle.vehicle_type = request.POST.get('vehicle_type')
        vehicle.driver_name = request.POST.get('driver_name')
        vehicle.status = request.POST.get('status', 'ACTIVE')
        vehicle.registration_number = request.POST.get('registration_number', '')
        vehicle.registration_expiry = request.POST.get('registration_expiry') or None
        vehicle.insurance_expiry = request.POST.get('insurance_expiry') or None
        vehicle.last_service_date = request.POST.get('last_service_date') or None
        vehicle.mileage = request.POST.get('mileage') or None
        vehicle.fuel_type = request.POST.get('fuel_type', 'Diesel')
        vehicle.capacity = request.POST.get('capacity') or None
        vehicle.notes = request.POST.get('notes', '')
        vehicle.purchase_date = request.POST.get('purchase_date') or None
        vehicle.purchase_price = request.POST.get('purchase_price') or None
        vehicle.current_value = request.POST.get('current_value') or None
        vehicle.is_active = request.POST.get('is_active') == 'on'
        
        # Check if vehicle number is unique (excluding current)
        if Vehicle.objects.filter(vehicle_number=vehicle.vehicle_number).exclude(id=vehicle.id).exists():
            messages.error(request, f'Vehicle "{vehicle.vehicle_number}" already exists.')
            return render(request, 'core/vehicle_form.html', {
                'action': 'Edit',
                'vehicle': vehicle,
                'vehicle_types': Vehicle.VEHICLE_TYPES,
                'fuel_types': Vehicle.FUEL_TYPES,
                'status_choices': Vehicle.STATUS_CHOICES,
            })
        
        vehicle.save()
        messages.success(request, f'Vehicle "{vehicle.vehicle_number}" updated successfully!')
        return redirect('core:vehicle_list')
    
    return render(request, 'core/vehicle_form.html', {
        'action': 'Edit',
        'vehicle': vehicle,
        'vehicle_types': Vehicle.VEHICLE_TYPES,
        'fuel_types': Vehicle.FUEL_TYPES,
        'status_choices': Vehicle.STATUS_CHOICES,
    })


@login_required
@permission_required('view_vehicles')
def vehicle_detail(request, vehicle_id):
    """View vehicle details with related data"""
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    
    # Get expenses
    expenses = Expense.objects.filter(vehicle=vehicle).order_by('-date')
    
    # Get sales
    sales = SalesBill.objects.filter(vehicle=vehicle).order_by('-date')
    
    # Get loads (VehicleStock)
    loads = VehicleStock.objects.filter(vehicle=vehicle).select_related('product')
    
    # Statistics
    total_expenses = vehicle.get_total_expenses()
    total_sales = vehicle.get_total_sales()
    net_performance = vehicle.get_net_performance()
    
    context = {
        'vehicle': vehicle,
        'expenses': expenses,
        'sales': sales,
        'loads': loads,
        'total_expenses': total_expenses,
        'total_sales': total_sales,
        'net_performance': net_performance,
        'expense_count': expenses.count(),
        'sale_count': sales.count(),
    }
    return render(request, 'core/vehicle_detail.html', context)


@login_required
@permission_required('manage_vehicles')
def vehicle_toggle_status(request, vehicle_id):
    """Activate or deactivate a vehicle"""
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    vehicle.is_active = not vehicle.is_active
    vehicle.status = 'ACTIVE' if vehicle.is_active else 'INACTIVE'
    vehicle.save()
    status = 'activated' if vehicle.is_active else 'deactivated'
    messages.success(request, f'Vehicle "{vehicle.vehicle_number}" has been {status}.')
    return redirect('core:vehicle_list')


@login_required
@permission_required('manage_vehicles')
def vehicle_delete(request, vehicle_id):
    """Delete a vehicle (with confirmation)"""
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    
    # Check if vehicle has related records
    has_expenses = Expense.objects.filter(vehicle=vehicle).exists()
    has_sales = SalesBill.objects.filter(vehicle=vehicle).exists()
    has_stock = VehicleStock.objects.filter(vehicle=vehicle).exists()
    
    if has_expenses or has_sales or has_stock:
        messages.error(request, f'Cannot delete "{vehicle.vehicle_number}" as it has related records.')
        return redirect('core:vehicle_list')
    
    if request.method == 'POST':
        vehicle.delete()
        messages.success(request, f'Vehicle "{vehicle.vehicle_number}" deleted successfully.')
        return redirect('core:vehicle_list')
    
    return render(request, 'core/vehicle_confirm_delete.html', {'vehicle': vehicle})


@login_required
@permission_required('manage_transfers')
def transfer_create(request):
    """Create a new stock transfer between vehicles"""
    vehicles = Vehicle.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True)
    
    if request.method == 'POST':
        source_id = request.POST.get('source_vehicle')
        dest_id = request.POST.get('destination_vehicle')
        product_id = request.POST.get('product')
        quantity = request.POST.get('quantity')
        reason = request.POST.get('reason', 'STOCK_BALANCE')
        notes = request.POST.get('notes', '')
        
        if not source_id or not dest_id or not product_id or not quantity:
            messages.error(request, 'Please fill in all required fields.')
            return render(request, 'core/transfer_create.html', {
                'vehicles': vehicles,
                'products': products,
                'reasons': StockTransfer.REASON_CHOICES,
            })
        
        if source_id == dest_id:
            messages.error(request, 'Source and destination vehicles cannot be the same.')
            return render(request, 'core/transfer_create.html', {
                'vehicles': vehicles,
                'products': products,
                'reasons': StockTransfer.REASON_CHOICES,
            })
        
        source = get_object_or_404(Vehicle, id=source_id)
        dest = get_object_or_404(Vehicle, id=dest_id)
        product = get_object_or_404(Product, id=product_id)
        quantity = Decimal(quantity)
        
        if quantity <= 0:
            messages.error(request, 'Quantity must be greater than zero.')
            return render(request, 'core/transfer_create.html', {
                'vehicles': vehicles,
                'products': products,
                'reasons': StockTransfer.REASON_CHOICES,
            })
        
        # Check if source vehicle has enough stock
        source_stock = VehicleStock.objects.filter(vehicle=source, product=product).first()
        if not source_stock or source_stock.quantity < quantity:
            messages.error(request, f'Source vehicle does not have enough {product.name}. Available: {source_stock.quantity if source_stock else 0}')
            return render(request, 'core/transfer_create.html', {
                'vehicles': vehicles,
                'products': products,
                'reasons': StockTransfer.REASON_CHOICES,
            })
        
        with transaction.atomic():
            # Deduct from source
            source_stock.quantity -= quantity
            source_stock.save()
            
            # Add to destination
            dest_stock, created = VehicleStock.objects.get_or_create(
                vehicle=dest,
                product=product
            )
            dest_stock.quantity += quantity
            dest_stock.save()
            
            # Create transfer record
            transfer = StockTransfer.objects.create(
                source_vehicle=source,
                destination_vehicle=dest,
                product=product,
                quantity=quantity,
                transferred_by=request.user,
                reason=reason,
                notes=notes,
            )
        
        messages.success(request, f'Successfully transferred {quantity} {product.unit} of {product.name} from {source.vehicle_number} to {dest.vehicle_number}.')
        return redirect('core:transfer_list')
    
    return render(request, 'core/transfer_create.html', {
        'vehicles': vehicles,
        'products': products,
        'reasons': StockTransfer.REASON_CHOICES,
    })


@login_required
@permission_required('view_transfers')
def transfer_list(request):
    """List all transfers with filters"""
    transfers = StockTransfer.objects.all().select_related('source_vehicle', 'destination_vehicle', 'product', 'transferred_by')
    
    search = request.GET.get('search', '')
    if search:
        transfers = transfers.filter(
            Q(source_vehicle__vehicle_number__icontains=search) |
            Q(destination_vehicle__vehicle_number__icontains=search) |
            Q(product__name__icontains=search)
        )
    
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if start_date:
        transfers = transfers.filter(transfer_date__date__gte=start_date)
    if end_date:
        transfers = transfers.filter(transfer_date__date__lte=end_date)
    
    source_id = request.GET.get('source_vehicle', '')
    if source_id and source_id.isdigit():
        transfers = transfers.filter(source_vehicle_id=int(source_id))
    
    dest_id = request.GET.get('destination_vehicle', '')
    if dest_id and dest_id.isdigit():
        transfers = transfers.filter(destination_vehicle_id=int(dest_id))
    
    vehicles = Vehicle.objects.filter(is_active=True)
    
    context = {
        'transfers': transfers,
        'vehicles': vehicles,
        'search': search,
        'start_date': start_date,
        'end_date': end_date,
        'selected_source': source_id,
        'selected_dest': dest_id,
        'total_transfers': transfers.count(),
    }
    return render(request, 'core/transfer_list.html', context)


@login_required
@permission_required('view_transfers')
def transfer_detail(request, transfer_id):
    transfer = get_object_or_404(StockTransfer, id=transfer_id)
    return render(request, 'core/transfer_detail.html', {'transfer': transfer})


@login_required
@permission_required('manage_transfers')
def transfer_delete(request, transfer_id):
    transfer = get_object_or_404(StockTransfer, id=transfer_id)
    # Allow deletion only within 1 hour? Or let admin decide.
    if request.method == 'POST':
        transfer.delete()
        messages.success(request, 'Transfer record deleted.')
        return redirect('core:transfer_list')
    return render(request, 'core/transfer_confirm_delete.html', {'transfer': transfer})


@login_required
@permission_required('view_vehicles')
def vehicle_stock(request):
    """View current stock on all vehicles"""
    vehicles = Vehicle.objects.filter(is_active=True)
    stock_data = []
    
    for vehicle in vehicles:
        stocks = VehicleStock.objects.filter(vehicle=vehicle).select_related('product')
        total_items = stocks.aggregate(total=Sum('quantity'))['total'] or 0
        stock_data.append({
            'vehicle': vehicle,
            'stocks': stocks,
            'total_items': total_items,
        })
    
    context = {
        'stock_data': stock_data,
    }
    return render(request, 'core/vehicle_stock.html', context)


@login_required
@permission_required('view_sales')
def sales_print(request, bill_id):
    """Print a sales bill as a printable invoice"""
    bill = get_object_or_404(SalesBill, id=bill_id)
    items = bill.items.all().select_related('product')
    payments = bill.payments.all()
    
    # Determine status
    if payments.exists():
        payment_types = [p.type for p in payments]
        if 'Credit' in payment_types and 'Cash' not in payment_types and 'Cheque' not in payment_types:
            status = 'CREDIT'
        elif 'Cash' in payment_types or 'Cheque' in payment_types:
            status = 'COMPLETE'
        else:
            status = 'PENDING'
    else:
        status = 'PENDING'
    
    context = {
        'bill': bill,
        'items': items,
        'payments': payments,
        'status': status,
    }
    return render(request, 'core/sales_print.html', context)


@login_required
@permission_required('create_sales')
def sales_edit(request, bill_id):
    """Edit a sales bill (only for DRAFT status)"""
    bill = get_object_or_404(SalesBill, id=bill_id)
    
    if bill.status != 'DRAFT':
        messages.error(request, f'Cannot edit a bill with status: {bill.get_status_display()}')
        return redirect('core:sales_list')
    
    # Get existing items
    items = bill.items.all().select_related('product')
    payments = bill.payments.all()
    
    # Redirect to sales page with bill data (pre-fill)
    # We'll pre-populate the sales form with existing data
    return redirect(f'/sales/?edit={bill.id}&invoice_no={bill.invoice_no}')


@login_required
@permission_required('delete_sales')
def sales_delete(request, bill_id):
    """Delete a sales bill and restore stock to vehicle"""
    bill = get_object_or_404(SalesBill, id=bill_id)
    invoice_no = bill.invoice_no
    vehicle = bill.vehicle
    
    with transaction.atomic():
        # 1. Restore stock to vehicle (for all items)
        for item in bill.items.all():
            try:
                vehicle_stock = VehicleStock.objects.get(vehicle=vehicle, product=item.product)
                vehicle_stock.quantity += abs(item.quantity)  # Always add back (even for returns)
                vehicle_stock.save()
            except VehicleStock.DoesNotExist:
                # If stock record doesn't exist, create it
                VehicleStock.objects.create(
                    vehicle=vehicle,
                    product=item.product,
                    quantity=abs(item.quantity)
                )
        
        # 2. Delete cheque records linked to this bill
        Cheque.objects.filter(sales_bill=bill).delete()
        
        # 3. Delete online payment records
        OnlinePayment.objects.filter(bill=bill).delete()
        
        # 4. Delete multi payment records
        MultiPayment.objects.filter(bill=bill).delete()
        
        # 5. Delete payments
        bill.payments.all().delete()
        
        # 6. Delete items
        bill.items.all().delete()
        
        # 7. Delete the bill
        bill.delete()
    
    messages.success(request, f'✅ Bill {invoice_no} deleted successfully! Stock restored to {vehicle.vehicle_number}.')
    return redirect('core:sales_list')


@login_required
def api_vehicle_stock(request):
    vehicle_id = request.GET.get('vehicle')
    if not vehicle_id or not vehicle_id.isdigit():
        return JsonResponse({'stocks': {}})
    try:
        vehicle = Vehicle.objects.get(id=vehicle_id)
        stocks = VehicleStock.objects.filter(vehicle=vehicle).values('product_id', 'quantity')
        stock_dict = {str(s['product_id']): float(s['quantity']) for s in stocks}
        return JsonResponse({'stocks': stock_dict})
    except Vehicle.DoesNotExist:
        return JsonResponse({'stocks': {}})


@login_required
@permission_required('create_sales')
def session_check(request):
    """Check if there's an open session for the user or vehicle"""
    vehicle_id = request.GET.get('vehicle')
    rep_id = request.GET.get('rep')
    
    response = {
        'has_open_session': False,
        'session': None,
        'conflict': None,
        'is_active_session': False,
    }
    
    # Check if there's an open session for this rep
    if rep_id and rep_id.isdigit():
        existing_session = DailySession.objects.filter(
            rep_id=int(rep_id),
            status='OPEN'
        ).first()
        if existing_session:
            response['has_open_session'] = True
            response['session'] = {
                'id': existing_session.id,
                'session_id': existing_session.session_id,
                'vehicle': existing_session.vehicle.vehicle_number,
                'rep': existing_session.rep.name,
                'started_at': existing_session.started_at.strftime('%Y-%m-%d %H:%M'),
                'total_bills': existing_session.total_bills,
            }
            response['conflict'] = f"Rep {existing_session.rep.name} already has an open session."
            return JsonResponse(response)
    
    # Check if there's an open session for this vehicle
    if vehicle_id and vehicle_id.isdigit():
        existing_session = DailySession.objects.filter(
            vehicle_id=int(vehicle_id),
            status='OPEN'
        ).first()
        if existing_session:
            response['has_open_session'] = True
            response['session'] = {
                'id': existing_session.id,
                'session_id': existing_session.session_id,
                'vehicle': existing_session.vehicle.vehicle_number,
                'rep': existing_session.rep.name,
                'started_at': existing_session.started_at.strftime('%Y-%m-%d %H:%M'),
                'total_bills': existing_session.total_bills,
            }
            response['conflict'] = f"Vehicle {existing_session.vehicle.vehicle_number} is already in use."
            return JsonResponse(response)
    
    return JsonResponse(response)


@login_required
@permission_required('create_sales')
def session_start(request):
    """Start a new daily session"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})
    
    vehicle_id = request.POST.get('vehicle')
    rep_id = request.POST.get('rep')
    
    if not vehicle_id or not rep_id:
        return JsonResponse({'success': False, 'error': 'Vehicle and Rep are required.'})
    
    # Check for conflicts
    existing_session = DailySession.objects.filter(
        rep_id=rep_id,
        status='OPEN'
    ).first()
    if existing_session:
        return JsonResponse({
            'success': False, 
            'error': f"Rep {existing_session.rep.name} already has an open session.",
            'session_id': existing_session.session_id
        })
    
    existing_session = DailySession.objects.filter(
        vehicle_id=vehicle_id,
        status='OPEN'
    ).first()
    if existing_session:
        return JsonResponse({
            'success': False, 
            'error': f"Vehicle {existing_session.vehicle.vehicle_number} is already in use.",
            'session_id': existing_session.session_id
        })
    
    # Create session
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    rep = get_object_or_404(Employee, id=rep_id)
    
    session = DailySession.objects.create(
        vehicle=vehicle,
        rep=rep,
        status='OPEN',
        started_by=request.user,
    )
    
    return JsonResponse({
        'success': True,
        'session_id': session.session_id,
        'session_id_display': session.session_id,
    })


@login_required
@permission_required('create_sales')
def session_summary(request):
    """Generate daily session summary report"""
    session_id = request.GET.get('session_id')
    if not session_id:
        messages.error(request, 'No session specified.')
        return redirect('core:sales')
    
    session = get_object_or_404(DailySession, id=session_id)
    
    # Only allow if status is OPEN or PENDING
    if session.status == 'COMPLETED':
        messages.warning(request, f'Session {session.session_id} is already completed.')
        return redirect('core:sales')
    
    # Get all pending items
    pending_bills = session.get_pending_bills().select_related('vehicle', 'rep')
    pending_expenses = session.get_pending_expenses().select_related('vehicle', 'employee')
    
    # Get completed bills (for reference)
    completed_bills = session.get_completed_bills()
    
    # Calculate totals
    total_sales = pending_bills.aggregate(total=Sum('net_total'))['total'] or 0
    total_foc = pending_bills.aggregate(total=Sum('foc_value'))['total'] or 0
    total_expenses = pending_expenses.aggregate(total=Sum('amount'))['total'] or 0
    net_collection = total_sales - total_expenses
    
    context = {
        'session': session,
        'pending_bills': pending_bills,
        'pending_expenses': pending_expenses,
        'completed_bills': completed_bills,
        'total_sales': total_sales,
        'total_foc': total_foc,
        'total_expenses': total_expenses,
        'net_collection': net_collection,
        'bill_count': pending_bills.count(),
        'expense_count': pending_expenses.count(),
        'today': date.today(),
    }
    return render(request, 'core/session_summary.html', context)


@login_required
@permission_required('create_sales')
def session_complete(request):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('core:sales')
    
    session_id = request.POST.get('session_id')
    if not session_id:
        messages.error(request, 'No session specified.')
        return redirect('core:sales')
    
    try:
        session = DailySession.objects.get(id=session_id)
    except DailySession.DoesNotExist:
        messages.error(request, f'Session not found (ID: {session_id}).')
        return redirect('core:sales')
    
    if session.status == 'COMPLETED':
        messages.warning(request, f'Session {session.session_id} is already completed.')
        return redirect('core:sales')
    
    if session.status == 'CANCELLED':
        messages.error(request, f'Session {session.session_id} has been cancelled.')
        return redirect('core:sales')
    
    try:
        with transaction.atomic():
            # 1. Complete all pending bills
            pending_bills = session.get_pending_bills()
            bill_count = pending_bills.count()
            for bill in pending_bills:
                bill.status = 'COMPLETED'
                bill.completed_by = request.user
                bill.completed_at = timezone.now()
                bill.save()
                print(f"✅ Bill {bill.invoice_no} completed.")  # Debug
            
            # 2. Mark all pending expenses as PAID
            pending_expenses = session.get_pending_expenses()
            expense_count = pending_expenses.count()
            for expense in pending_expenses:
                expense.status = 'PAID'
                expense.paid_at = timezone.now()
                expense.save()
                print(f"✅ Expense {expense.id} paid.")  # Debug
            
            # 3. Update session
            session.status = 'COMPLETED'
            session.completed_at = timezone.now()
            session.completed_by = request.user
            session.update_summary()
            session.save()
            print(f"✅ Session {session.session_id} completed.")  # Debug
        
        messages.success(
            request, 
            f'✅ Session {session.session_id} completed! '
            f'{bill_count} bills and {expense_count} expenses finalized.'
        )
    except Exception as e:
        messages.error(request, f'❌ Error completing session: {str(e)}')
        print(f"❌ Error: {e}")  # Debug
        return redirect('core:sales')
    
    return redirect('core:sales')


@login_required
@permission_required('create_sales')
def pay_credit(request, bill_id):
    """Process partial/full payment for a credit bill"""
    bill = get_object_or_404(SalesBill, id=bill_id)

    paid_amount = bill.payments.exclude(type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
    outstanding = bill.net_total - paid_amount

    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', '0'))
        payment_method = request.POST.get('payment_method', 'Cash')
        notes = request.POST.get('notes', '')
        
        # Calculate outstanding (exclude Credit payments)
        paid_amount = bill.payments.exclude(type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        outstanding = bill.net_total - paid_amount
        
        # Validate
        if amount <= 0:
            messages.error(request, 'Amount must be greater than zero.')
            return redirect('core:credit_list')
        
        if amount > outstanding:
            messages.error(request, f'Payment amount ({amount}) exceeds outstanding balance ({outstanding}).')
            return redirect('core:credit_list')
        
        with transaction.atomic():
            # Create payment
            Payment.objects.create(
                bill=bill,
                type=payment_method,
                amount=amount
            )
            
            # If cheque, create cheque record
            if payment_method == 'Cheque':
                cheque_no = request.POST.get('cheque_no', '')
                cheque_date = request.POST.get('cheque_date', '')
                bank_id = request.POST.get('bank_id', '')
                if cheque_no and cheque_date and bank_id:
                    bank = get_object_or_404(Bank, id=bank_id)
                    Cheque.objects.create(
                        cheque_no=cheque_no,
                        bank=bank,
                        cheque_date=cheque_date,
                        amount=amount,
                        customer_name=bill.shop_name or 'N/A',
                        sales_bill=bill,
                        status='PENDING',
                        notes=f"Payment for credit bill: {bill.invoice_no}"
                    )
        
        messages.success(request, f'✅ Payment of Rs {amount} recorded for bill {bill.invoice_no}.')
        return redirect('core:credit_list')
    
    # GET request: show payment form
    context = {
        'bill': bill,
        'paid_amount': paid_amount,
        'outstanding': outstanding,
    }
    return render(request, 'core/pay_credit.html', context)


@login_required
@permission_required('view_reports')
def discount_history_report(request):
    from datetime import date
    start_date = request.GET.get('start_date', date.today().strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', date.today().strftime('%Y-%m-%d'))

    bills = SalesBill.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        status='COMPLETED'
    ).select_related('vehicle', 'rep')

    total_bill_discounts = 0
    total_product_discounts = 0
    discount_data = []

    for bill in bills:
        bill_discount = bill.bill_discount_amount or 0
        product_discounts = 0
        for item in bill.items.all():
            if item.discount_value > 0:
                product_discounts += (item.rate - item.discounted_rate) * item.quantity

        if bill_discount > 0 or product_discounts > 0:
            total_bill_discounts += bill_discount
            total_product_discounts += product_discounts
            discount_data.append({
                'bill': bill,
                'bill_discount': bill_discount,
                'product_discounts': product_discounts,
            })

    context = {
        'discount_data': discount_data,
        'start_date': start_date,
        'end_date': end_date,
        'total_bill_discounts': total_bill_discounts,
        'total_product_discounts': total_product_discounts,
        'total_discounts': total_bill_discounts + total_product_discounts,
        'count': len(discount_data),
    }
    return render(request, 'core/discount_history_report.html', context)


@login_required
@permission_required('view_reports')
def day_selling_report(request):
    """Day Selling Report with filterable statuses and Excel export"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    vehicle_id = request.GET.get('vehicle', '')
    
    # Parse dates
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    # Get status filters from request (checkboxes)
    show_complete = request.GET.get('show_complete', 'on') == 'on'
    show_credit = request.GET.get('show_credit', 'on') == 'on'
    show_credit_payments = request.GET.get('show_credit_payments', 'on') == 'on'
    show_returns = request.GET.get('show_returns', 'on') == 'on'
    show_expenses = request.GET.get('show_expenses', 'on') == 'on'
    
    # Filter bills
    bills = SalesBill.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj,
        status='COMPLETED'
    ).select_related('vehicle', 'rep')
    
    if vehicle_id and vehicle_id.isdigit():
        bills = bills.filter(vehicle_id=int(vehicle_id))
    
    # ====== COMPUTE STATUS FOR EACH BILL ======
    for bill in bills:
        paid_amount = bill.payments.exclude(type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        if paid_amount >= bill.net_total:
            bill.display_status = 'COMPLETE'
            bill.display_status_color = '#28a745'
        else:
            bill.display_status = 'CREDIT'
            bill.display_status_color = '#fd7e14'
        bill.paid_amount = paid_amount
    
    # ====== CALCULATIONS (with filters applied) ======
    complete_sales = Decimal('0')
    complete_bills = []
    if show_complete:
        for bill in bills:
            if bill.display_status == 'COMPLETE':
                complete_sales += bill.net_total
                complete_bills.append(bill)
    
    # Credit Sales
    credit_bill_ids = Payment.objects.filter(bill__in=bills, type='Credit').values_list('bill_id', flat=True).distinct()
    credit_bills = SalesBill.objects.filter(id__in=credit_bill_ids)
    credit_sales = credit_bills.aggregate(total=Sum('net_total'))['total'] or Decimal('0') if show_credit else Decimal('0')
    
    # Credit Payments
    credit_payments = Decimal('0')
    if show_credit_payments:
        credit_payments = Payment.objects.filter(
            bill__in=credit_bill_ids,
            type__in=['Cash', 'Cheque', 'Online']
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Returns
    returns_total = Decimal('0')
    if show_returns:
        return_items = SalesItem.objects.filter(bill__in=bills, quantity__lt=0)
        returns_total = abs(return_items.aggregate(total=Sum('total'))['total'] or Decimal('0'))
    
    # Expenses
    expense_total = Decimal('0')
    if show_expenses:
        expenses = Expense.objects.filter(date__gte=start_date_obj, date__lte=end_date_obj)
        if vehicle_id and vehicle_id.isdigit():
            expenses = expenses.filter(vehicle_id=int(vehicle_id))
        expense_total = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # ====== EXCEL EXPORT ======
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Day Selling Report"
        
        # Headers
        headers = ['Metric', 'Amount (Rs)']
        ws.append(headers)
        for col in range(1, 3):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Data rows
        ws.append(['Complete Sales', float(complete_sales)])
        ws.append(['Credit Sales', float(credit_sales)])
        ws.append(['Credit Payments', float(credit_payments)])
        ws.append(['Returns', float(returns_total)])
        ws.append(['Expenses', float(expense_total)])
        
        # Add bill details
        ws.append([])
        ws.append(['Invoice', 'Customer', 'Amount (Rs)', 'Date', 'Status'])
        for col in range(1, 6):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        for bill in bills:
            ws.append([
                bill.invoice_no,
                bill.shop_name,
                float(bill.net_total),
                bill.date.strftime('%Y-%m-%d'),
                bill.display_status
            ])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Day_Selling_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    vehicles = Vehicle.objects.filter(is_active=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'selected_vehicle': vehicle_id,
        'vehicles': vehicles,
        'complete_sales': complete_sales,
        'credit_sales': credit_sales,
        'credit_payments': credit_payments,
        'returns_total': returns_total,
        'expense_total': expense_total,
        'complete_bills_count': len(complete_bills),
        'credit_bills_count': credit_bills.count(),
        'show_complete': show_complete,
        'show_credit': show_credit,
        'show_credit_payments': show_credit_payments,
        'show_returns': show_returns,
        'show_expenses': show_expenses,
        'bills': bills,
    }
    
    return render(request, 'core/day_selling_report.html', context)


@login_required
@permission_required('view_reports')
def sold_items_report(request):
    """Sold Items Report with optional grouping by vehicle"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    vehicle_id = request.GET.get('vehicle', '')
    product_id = request.GET.get('product', '')
    category_id = request.GET.get('category', '')
    group_by_vehicle = request.GET.get('group_by_vehicle', 'off') == 'on'
    
    # Parse dates
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today.replace(day=1)
        end_date_obj = today
    
    # Base query: SalesBill completed, date range
    bills = SalesBill.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj,
        status='COMPLETED'
    )
    
    if vehicle_id and vehicle_id.isdigit():
        bills = bills.filter(vehicle_id=int(vehicle_id))
    
    # SalesItems from those bills
    items = SalesItem.objects.filter(bill__in=bills).select_related('product', 'bill__vehicle')
    
    # Product filter
    if product_id and product_id.isdigit():
        items = items.filter(product_id=int(product_id))
    
    # Category filter (via product)
    if category_id and category_id.isdigit():
        items = items.filter(product__category_id=int(category_id))
    
    # ====== Aggregation ======
    # Total summary (across all items)
    total_qty = items.filter(is_foc=False).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    total_amount = items.filter(is_foc=False).aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_foc_qty = items.filter(is_foc=True).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    total_foc_value = items.filter(is_foc=True).aggregate(total=Sum('total'))['total'] or Decimal('0')
    invoice_count = bills.count()
    unique_products = items.values('product_id').distinct().count()
    
    # ====== Group by Vehicle ======
    if group_by_vehicle:
        # Get all vehicles involved
        vehicle_ids = bills.values_list('vehicle_id', flat=True).distinct()
        vehicles = Vehicle.objects.filter(id__in=vehicle_ids)
        grouped_data = []
        
        for vehicle in vehicles:
            vehicle_items = items.filter(bill__vehicle=vehicle)
            # Group by product within this vehicle
            product_rows = []
            # Use values + annotate
            product_agg = vehicle_items.values('product_id', 'product__name', 'product__unit', 'product__category__name').annotate(
                qty_sold=Sum('quantity', filter=Q(is_foc=False)),
                amount=Sum('total', filter=Q(is_foc=False)),
                foc_qty=Sum('quantity', filter=Q(is_foc=True)),
                foc_value=Sum('total', filter=Q(is_foc=True)),
            )
            for row in product_agg:
                qty_sold = row.get('qty_sold') or Decimal('0')
                amount = row.get('amount') or Decimal('0')
                foc_qty = row.get('foc_qty') or Decimal('0')
                foc_value = row.get('foc_value') or Decimal('0')
                avg_rate = (amount / qty_sold) if qty_sold > 0 else Decimal('0')
                product_rows.append({
                    'name': row['product__name'],
                    'category': row['product__category__name'] or 'Uncategorized',
                    'unit': row['product__unit'],
                    'qty_sold': qty_sold,
                    'amount': amount,
                    'foc_qty': foc_qty,
                    'foc_value': foc_value,
                    'avg_rate': avg_rate,
                })
            # Sort by amount descending
            product_rows.sort(key=lambda x: x['amount'], reverse=True)
            grouped_data.append({
                'vehicle': vehicle,
                'rows': product_rows,
                'vehicle_total_qty': sum(r['qty_sold'] for r in product_rows),
                'vehicle_total_amount': sum(r['amount'] for r in product_rows),
            })
    else:
        # Flat list (no grouping)
        product_agg = items.values('product_id', 'product__name', 'product__unit', 'product__category__name').annotate(
            qty_sold=Sum('quantity', filter=Q(is_foc=False)),
            amount=Sum('total', filter=Q(is_foc=False)),
            foc_qty=Sum('quantity', filter=Q(is_foc=True)),
            foc_value=Sum('total', filter=Q(is_foc=True)),
        )
        flat_rows = []
        for row in product_agg:
            qty_sold = row.get('qty_sold') or Decimal('0')
            amount = row.get('amount') or Decimal('0')
            foc_qty = row.get('foc_qty') or Decimal('0')
            foc_value = row.get('foc_value') or Decimal('0')
            avg_rate = (amount / qty_sold) if qty_sold > 0 else Decimal('0')
            flat_rows.append({
                'name': row['product__name'],
                'category': row['product__category__name'] or 'Uncategorized',
                'unit': row['product__unit'],
                'qty_sold': qty_sold,
                'amount': amount,
                'foc_qty': foc_qty,
                'foc_value': foc_value,
                'avg_rate': avg_rate,
            })
        flat_rows.sort(key=lambda x: x['amount'], reverse=True)
        grouped_data = [{'vehicle': None, 'rows': flat_rows, 'vehicle_total_qty': total_qty, 'vehicle_total_amount': total_amount}]
    
    # ====== Excel Export ======
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Sold Items Report"
        
        # Headers
        headers = ['Product', 'Category', 'Unit', 'Qty Sold', 'Amount (Rs)', 'FOC Qty', 'FOC Value (Rs)', 'Avg Rate (Rs)']
        ws.append(headers)
        for col in range(1, 9):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Data rows
        if group_by_vehicle:
            for group in grouped_data:
                # Add vehicle header
                ws.append([f"Vehicle: {group['vehicle'].vehicle_number} - {group['vehicle'].driver_name}"])
                for row in group['rows']:
                    ws.append([
                        row['name'],
                        row['category'],
                        row['unit'],
                        float(row['qty_sold']),
                        float(row['amount']),
                        float(row['foc_qty']),
                        float(row['foc_value']),
                        float(row['avg_rate']),
                    ])
                ws.append([])  # blank row between vehicles
        else:
            for row in flat_rows:
                ws.append([
                    row['name'],
                    row['category'],
                    row['unit'],
                    float(row['qty_sold']),
                    float(row['amount']),
                    float(row['foc_qty']),
                    float(row['foc_value']),
                    float(row['avg_rate']),
                ])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Sold_Items_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    # Context for template
    vehicles = Vehicle.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True)
    categories = Category.objects.filter(is_active=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'selected_vehicle': vehicle_id,
        'selected_product': product_id,
        'selected_category': category_id,
        'group_by_vehicle': group_by_vehicle,
        'vehicles': vehicles,
        'products': products,
        'categories': categories,
        'grouped_data': grouped_data,
        'total_qty': total_qty,
        'total_amount': total_amount,
        'total_foc_qty': total_foc_qty,
        'total_foc_value': total_foc_value,
        'invoice_count': invoice_count,
        'unique_products': unique_products,
    }
    return render(request, 'core/sold_items_report.html', context)


@login_required
@permission_required('view_reports')
def free_issue_report(request):
    """Free Issue (FOC) Report – shows all FOC items given away"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    
    # Parse dates
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    # Get all FOC items from bills in date range
    foc_items = SalesItem.objects.filter(
        is_foc=True,
        bill__status='COMPLETED',
        bill__date__gte=start_date_obj,
        bill__date__lte=end_date_obj
    ).select_related('product', 'bill', 'bill__vehicle', 'bill__rep')
    
    # Calculate summary
    total_foc_qty = foc_items.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    total_foc_value = foc_items.aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_invoices = foc_items.values('bill_id').distinct().count()
    unique_products = foc_items.values('product_id').distinct().count()
    
    # Group by product for detailed table
    product_agg = foc_items.values('product_id', 'product__name', 'product__unit').annotate(
        total_qty=Sum('quantity'),
        total_value=Sum('total')
    ).order_by('-total_value')
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Free Issue Report"
        
        headers = ['Date', 'Invoice', 'Customer', 'Vehicle', 'Rep', 'Product', 'Unit', 'Qty', 'Rate', 'Total Value']
        ws.append(headers)
        for col in range(1, 11):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        for item in foc_items:
            ws.append([
                item.bill.date.strftime('%Y-%m-%d'),
                item.bill.invoice_no,
                item.bill.shop_name or 'N/A',
                item.bill.vehicle.vehicle_number if item.bill.vehicle else 'N/A',
                item.bill.rep.name if item.bill.rep else 'N/A',
                item.product.name,
                item.product.unit,
                float(item.quantity),
                float(item.rate),
                float(item.total),
            ])
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Free_Issue_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'foc_items': foc_items,
        'total_foc_qty': total_foc_qty,
        'total_foc_value': total_foc_value,
        'total_invoices': total_invoices,
        'unique_products': unique_products,
        'product_agg': product_agg,
        'today': today,
    }
    return render(request, 'core/free_issue_report.html', context)


@login_required
@permission_required('view_reports')
def discount_report(request):
    """Discount Report – shows all bill-level discounts applied"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    # Get all completed bills with discounts (bill_discount_amount > 0)
    bills = SalesBill.objects.filter(
        status='COMPLETED',
        date__gte=start_date_obj,
        date__lte=end_date_obj,
        bill_discount_amount__gt=0
    ).select_related('vehicle', 'rep')
    
    # Summary
    total_discount_amount = bills.aggregate(total=Sum('bill_discount_amount'))['total'] or Decimal('0')
    total_bills = bills.count()
    total_value_after_discount = bills.aggregate(total=Sum('net_total'))['total'] or Decimal('0')
    total_original_value = total_value_after_discount + total_discount_amount
    
    # Group by discount type (Percentage vs Fixed)
    percent_count = bills.filter(bill_discount_type='PERCENTAGE').count()
    fixed_count = bills.filter(bill_discount_type='FIXED').count()
    percent_total = bills.filter(bill_discount_type='PERCENTAGE').aggregate(total=Sum('bill_discount_amount'))['total'] or Decimal('0')
    fixed_total = bills.filter(bill_discount_type='FIXED').aggregate(total=Sum('bill_discount_amount'))['total'] or Decimal('0')
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Discount Report"
        
        headers = ['Invoice', 'Customer', 'Vehicle', 'Date', 'Discount Type', 'Discount Value', 'Discount Amount', 'Net Total']
        ws.append(headers)
        for col in range(1, 9):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        for bill in bills:
            ws.append([
                bill.invoice_no,
                bill.shop_name or 'N/A',
                bill.vehicle.vehicle_number if bill.vehicle else 'N/A',
                bill.date.strftime('%Y-%m-%d'),
                bill.get_bill_discount_type_display() or 'None',
                float(bill.bill_discount_value),
                float(bill.bill_discount_amount),
                float(bill.net_total),
            ])
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Discount_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'bills': bills,
        'total_discount_amount': total_discount_amount,
        'total_bills': total_bills,
        'total_value_after_discount': total_value_after_discount,
        'total_original_value': total_original_value,
        'percent_count': percent_count,
        'fixed_count': fixed_count,
        'percent_total': percent_total,
        'fixed_total': fixed_total,
    }
    return render(request, 'core/discount_report.html', context)


@login_required
@permission_required('view_reports')
def expense_by_vehicle_report(request):
    """Expense by Vehicle & Category report with filters"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    category = request.GET.get('category', '')
    user_id = request.GET.get('user', '')
    vehicle_id = request.GET.get('vehicle', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    # Base queryset
    expenses = Expense.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )
    
    # Apply filters
    if category:
        expenses = expenses.filter(category=category)
    if user_id and user_id.isdigit():
        expenses = expenses.filter(employee_id=int(user_id))
    if vehicle_id and vehicle_id.isdigit():
        expenses = expenses.filter(vehicle_id=int(vehicle_id))
    
    # Group by vehicle and category
    # Get all distinct vehicles in the filtered expenses (or all active if none)
    if expenses.exists():
        vehicle_ids = expenses.values_list('vehicle_id', flat=True).distinct()
        vehicles = Vehicle.objects.filter(id__in=vehicle_ids)
    else:
        vehicles = Vehicle.objects.filter(is_active=True)
    
    # Category list from model choices
    category_choices = Expense.CATEGORY_CHOICES
    category_keys = [c[0] for c in category_choices]
    category_labels = {c[0]: c[1] for c in category_choices}
    
    # Build pivot table
    pivot_data = []
    grand_total = Decimal('0')
    category_totals = {cat: Decimal('0') for cat in category_keys}
    
    for vehicle in vehicles:
        row = {
            'vehicle': vehicle,
            'categories': {},
            'total': Decimal('0'),
        }
        # Get expenses for this vehicle
        veh_expenses = expenses.filter(vehicle=vehicle)
        for cat in category_keys:
            cat_total = veh_expenses.filter(category=cat).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            row['categories'][cat] = cat_total
            row['total'] += cat_total
            category_totals[cat] += cat_total
        grand_total += row['total']
        pivot_data.append(row)
    
    # Sort by total descending
    pivot_data.sort(key=lambda x: x['total'], reverse=True)
    
    # Get options for filters
    categories = category_choices
    users = Employee.objects.filter(is_active=True)
    vehicles_list = Vehicle.objects.filter(is_active=True)
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense by Vehicle & Category"
        
        # Headers: Vehicle, Driver, then each category, then Total
        headers = ['Vehicle', 'Driver']
        headers.extend([category_labels[cat] for cat in category_keys])
        headers.append('Total')
        ws.append(headers)
        for col in range(1, len(headers)+1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Data rows
        for row in pivot_data:
            row_data = [
                row['vehicle'].vehicle_number,
                row['vehicle'].driver_name,
            ]
            for cat in category_keys:
                row_data.append(float(row['categories'][cat]))
            row_data.append(float(row['total']))
            ws.append(row_data)
        
        # Add category totals row
        total_row = ['Grand Total', '']
        for cat in category_keys:
            total_row.append(float(category_totals[cat]))
        total_row.append(float(grand_total))
        ws.append(total_row)
        # Make grand total row bold
        for col in range(1, len(total_row)+1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Expense_by_Vehicle_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'selected_category': category,
        'selected_user': user_id,
        'selected_vehicle': vehicle_id,
        'categories': categories,
        'users': users,
        'vehicles': vehicles_list,
        'pivot_data': pivot_data,
        'category_keys': category_keys,
        'category_labels': category_labels,
        'grand_total': grand_total,
        'category_totals': category_totals,
        'total_vehicles': len(pivot_data),
    }
    return render(request, 'core/expense_by_vehicle_report.html', context)


@login_required
@permission_required('view_reports')
def sales_by_rep_report(request):
    """Sales by Rep Report with date range and rep filters"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    rep_id = request.GET.get('rep', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    # Filter bills by date range
    bills = SalesBill.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj,
        status='COMPLETED'
    ).select_related('rep', 'vehicle')
    
    # Filter by rep
    if rep_id and rep_id.isdigit():
        bills = bills.filter(rep_id=int(rep_id))
    
    # ====== Group by Rep ======
    rep_data = {}
    
    for bill in bills:
        rep_name = bill.rep.name if bill.rep else 'Unassigned'
        if rep_name not in rep_data:
            rep_data[rep_name] = {
                'rep': bill.rep,
                'total_sales': Decimal('0'),
                'total_foc': Decimal('0'),
                'bill_count': 0,
                'bills': [],
            }
        rep_data[rep_name]['total_sales'] += bill.net_total
        rep_data[rep_name]['total_foc'] += bill.foc_value or Decimal('0')
        rep_data[rep_name]['bill_count'] += 1
        rep_data[rep_name]['bills'].append(bill)
    
    # Sort by total sales descending
    rep_list = sorted(rep_data.values(), key=lambda x: x['total_sales'], reverse=True)
    
    # Overall totals
    total_sales = sum(r['total_sales'] for r in rep_list)
    total_foc = sum(r['total_foc'] for r in rep_list)
    total_bills = sum(r['bill_count'] for r in rep_list)
    
    # Get reps for filter dropdown
    reps = Employee.objects.filter(position='Rep', is_active=True)
    
    # ====== Excel Export ======
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales by Rep Report"
        
        # Headers
        headers = ['Rep', 'Total Sales (Rs)', 'FOC Value (Rs)', 'Bill Count']
        ws.append(headers)
        for col in range(1, 5):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Data rows
        for rep in rep_list:
            ws.append([
                rep_name,
                float(rep['total_sales']),
                float(rep['total_foc']),
                rep['bill_count'],
            ])
        
        # Add bill details
        ws.append([])
        ws.append(['Invoice', 'Customer', 'Vehicle', 'Amount (Rs)', 'Date'])
        for col in range(1, 6):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        
        for rep in rep_list:
            ws.append([f"--- {rep['rep'].name if rep['rep'] else 'Unassigned'} ---"])
            for bill in rep['bills']:
                ws.append([
                    bill.invoice_no,
                    bill.shop_name or 'N/A',
                    bill.vehicle.vehicle_number if bill.vehicle else 'N/A',
                    float(bill.net_total),
                    bill.date.strftime('%Y-%m-%d'),
                ])
            ws.append([])  # blank row between reps
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Sales_by_Rep_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'selected_rep': rep_id,
        'reps': reps,
        'rep_list': rep_list,
        'total_sales': total_sales,
        'total_foc': total_foc,
        'total_bills': total_bills,
        'today': today,
    }
    
    return render(request, 'core/sales_by_rep_report.html', context)


@login_required
@permission_required('view_reports')
def return_report(request):
    """Return Report with date range and return type filters"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    return_type = request.GET.get('return_type', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    # Get all return items (negative quantities)
    # A return is identified by quantity < 0 in SalesItem
    return_items = SalesItem.objects.filter(
        quantity__lt=0,
        bill__status='COMPLETED',
        bill__date__gte=start_date_obj,
        bill__date__lte=end_date_obj
    ).select_related('product', 'bill', 'bill__vehicle', 'bill__rep', 'bill__customer')
    
    # Filter by return type (if you have a return_reason field on SalesItem or SalesBill)
    # For this, we assume you have a return_reason field on SalesItem or SalesBill
    # If not, we'll filter by bill.return_reason (which we added earlier)
    
    if return_type:
        # If you store return_reason on the SalesItem
        return_items = return_items.filter(return_reason=return_type)
        # Alternatively, if stored on SalesBill:
        # return_items = return_items.filter(bill__return_reason=return_type)
    
    # Calculate summary
    total_return_qty = abs(return_items.aggregate(total=Sum('quantity'))['total'] or Decimal('0'))
    total_return_value = return_items.aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_return_value = abs(total_return_value)
    total_invoices = return_items.values('bill_id').distinct().count()
    unique_products = return_items.values('product_id').distinct().count()
    
    # Group by return reason (if you have it)
    reason_breakdown = {}
    for item in return_items:
        reason = item.return_reason or 'OTHER'
        if reason not in reason_breakdown:
            reason_breakdown[reason] = {
                'count': 0,
                'value': Decimal('0'),
            }
        reason_breakdown[reason]['count'] += 1
        reason_breakdown[reason]['value'] += abs(item.total)
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Return Report"
        
        headers = ['Date', 'Invoice', 'Customer', 'Vehicle', 'Rep', 'Product', 'Qty Returned', 'Rate', 'Total Value', 'Reason']
        ws.append(headers)
        for col in range(1, 11):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        for item in return_items:
            ws.append([
                item.bill.date.strftime('%Y-%m-%d'),
                item.bill.invoice_no,
                item.bill.shop_name or 'N/A',
                item.bill.vehicle.vehicle_number if item.bill.vehicle else 'N/A',
                item.bill.rep.name if item.bill.rep else 'N/A',
                item.product.name,
                float(abs(item.quantity)),
                float(item.rate),
                float(abs(item.total)),
                item.return_reason or 'Other',
            ])
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Return_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    # Return type choices for filter (if you have a model with choices)
    return_types = [
        ('DAMAGED', 'Damaged'),
        ('EXPIRED', 'Expired'),
        ('OTHER', 'Other'),
    ]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'selected_return_type': return_type,
        'return_types': return_types,
        'return_items': return_items,
        'total_return_qty': total_return_qty,
        'total_return_value': total_return_value,
        'total_invoices': total_invoices,
        'unique_products': unique_products,
        'reason_breakdown': reason_breakdown,
        'today': today,
    }
    
    return render(request, 'core/return_report.html', context)


@login_required
@permission_required('view_reports')
def cheque_report(request):
    """Cheque Report with date range and status filters"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    status_filter = request.GET.get('status', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    # Get all cheques with filtering
    cheques = Cheque.objects.filter(
        created_at__date__gte=start_date_obj,
        created_at__date__lte=end_date_obj
    ).select_related('bank', 'sales_bill', 'sales_bill__customer')
    
    if status_filter:
        cheques = cheques.filter(status=status_filter)
    
    # Calculate summary by status
    total_cheques = cheques.count()
    total_amount = cheques.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Status breakdown
    pending_cheques = cheques.filter(status='PENDING')
    deposited_cheques = cheques.filter(status='DEPOSITED')
    cleared_cheques = cheques.filter(status='CLEARED')
    bounced_cheques = cheques.filter(status='BOUNCED')
    
    pending_count = pending_cheques.count()
    pending_amount = pending_cheques.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    deposited_count = deposited_cheques.count()
    deposited_amount = deposited_cheques.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    cleared_count = cleared_cheques.count()
    cleared_amount = cleared_cheques.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    bounced_count = bounced_cheques.count()
    bounced_amount = bounced_cheques.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Additional summary
    total_received = cleared_count + deposited_count + bounced_count + pending_count
    total_received_amount = cleared_amount + deposited_amount + bounced_amount + pending_amount
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Cheque Report"
        
        # Status Summary
        ws.append(['Cheque Status Summary'])
        ws.append(['Status', 'Count', 'Total Amount (Rs)'])
        for col in range(1, 4):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        ws.append(['Pending', pending_count, float(pending_amount)])
        ws.append(['Deposited', deposited_count, float(deposited_amount)])
        ws.append(['Cleared', cleared_count, float(cleared_amount)])
        ws.append(['Bounced', bounced_count, float(bounced_amount)])
        ws.append(['Total', total_cheques, float(total_amount)])
        
        # Cheque Details
        ws.append([])
        ws.append(['Cheque Details'])
        headers = ['Cheque No', 'Bank', 'Customer', 'Amount (Rs)', 'Cheque Date', 'Status', 'Invoice']
        ws.append(headers)
        for col in range(1, 8):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        
        for cheque in cheques:
            ws.append([
                cheque.cheque_no,
                cheque.bank.name,
                cheque.customer_name,
                float(cheque.amount),
                cheque.cheque_date.strftime('%Y-%m-%d'),
                cheque.get_status_display(),
                cheque.sales_bill.invoice_no if cheque.sales_bill else 'N/A',
            ])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Cheque_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    # Status choices for filter
    status_choices = Cheque.STATUS_CHOICES
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'selected_status': status_filter,
        'status_choices': status_choices,
        'cheques': cheques,
        'total_cheques': total_cheques,
        'total_amount': total_amount,
        'pending_count': pending_count,
        'pending_amount': pending_amount,
        'deposited_count': deposited_count,
        'deposited_amount': deposited_amount,
        'cleared_count': cleared_count,
        'cleared_amount': cleared_amount,
        'bounced_count': bounced_count,
        'bounced_amount': bounced_amount,
        'total_received': total_received,
        'total_received_amount': total_received_amount,
        'today': today,
    }
    
    return render(request, 'core/cheque_report.html', context)


@login_required
@permission_required('view_reports')
def online_payment_report(request):
    """Online Payment Report with date range and method filters"""
    
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    method_filter = request.GET.get('method', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    # Get all online payments
    online_payments = OnlinePayment.objects.filter(
        created_at__date__gte=start_date_obj,
        created_at__date__lte=end_date_obj
    ).select_related('bill', 'bill__customer', 'bill__vehicle', 'bill__rep')
    
    if method_filter:
        online_payments = online_payments.filter(payment_method=method_filter)
    
    # Calculate summary
    total_payments = online_payments.count()
    total_amount = online_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Method breakdown
    method_breakdown = online_payments.values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-total')
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Online Payment Report"
        
        # Summary
        ws.append(['Online Payment Report'])
        ws.append([])
        ws.append(['Total Payments', total_payments])
        ws.append(['Total Amount', float(total_amount)])
        ws.append([])
        ws.append(['Payment Method', 'Count', 'Total Amount (Rs)'])
        for col in range(1, 4):
            ws.cell(row=5, column=col).font = Font(bold=True)
        
        row = 6
        for item in method_breakdown:
            ws.append([item['payment_method'], item['count'], float(item['total'])])
            row += 1
        
        ws.append([])
        ws.append(['Transaction Details'])
        headers = ['Date', 'Invoice', 'Customer', 'Vehicle', 'Rep', 'Payment Method', 'Reference', 'Amount (Rs)']
        ws.append(headers)
        for col in range(1, 9):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        
        for payment in online_payments:
            ws.append([
                payment.created_at.strftime('%Y-%m-%d %H:%M'),
                payment.bill.invoice_no if payment.bill else 'N/A',
                payment.bill.shop_name if payment.bill else 'N/A',
                payment.bill.vehicle.vehicle_number if payment.bill and payment.bill.vehicle else 'N/A',
                payment.bill.rep.name if payment.bill and payment.bill.rep else 'N/A',
                payment.payment_method,
                payment.reference_no or 'N/A',
                float(payment.amount),
            ])
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Online_Payment_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    # Payment methods for filter
    payment_methods = OnlinePayment.PAYMENT_METHODS if hasattr(OnlinePayment, 'PAYMENT_METHODS') else [
        ('Bank Transfer', 'Bank Transfer'),
        ('Mobile Wallet', 'Mobile Wallet'),
        ('Card', 'Card'),
        ('Other', 'Other'),
    ]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'selected_method': method_filter,
        'payment_methods': payment_methods,
        'online_payments': online_payments,
        'total_payments': total_payments,
        'total_amount': total_amount,
        'method_breakdown': method_breakdown,
        'today': today,
    }
    
    return render(request, 'core/online_payment_report.html', context)


@login_required
@permission_required('view_reports')
def inventory_by_location_report(request):
    """Inventory by Location Report (Warehouse, Vehicles, or Both)"""
    
    today = date.today()
    location_filter = request.GET.get('location', 'all')  # all, warehouse, vehicles
    category_id = request.GET.get('category', '')
    
    # Get all active products
    products = Product.objects.filter(is_active=True)
    
    # Filter by category if selected
    if category_id and category_id.isdigit():
        products = products.filter(category_id=int(category_id))
    
    # Prepare stock data
    stock_data = []
    total_warehouse_qty = Decimal('0')
    total_vehicle_qty = Decimal('0')
    total_stock_qty = Decimal('0')
    total_warehouse_value = Decimal('0')
    total_vehicle_value = Decimal('0')
    total_stock_value = Decimal('0')
    
    for product in products:
        warehouse_qty = product.get_warehouse_stock()  # from Product model method
        vehicle_qty = product.get_vehicle_stock()
        
        # Apply location filter
        if location_filter == 'warehouse':
            if warehouse_qty == 0:
                continue
            display_warehouse = warehouse_qty
            display_vehicle = Decimal('0')
            display_total = warehouse_qty
        elif location_filter == 'vehicles':
            if vehicle_qty == 0:
                continue
            display_warehouse = Decimal('0')
            display_vehicle = vehicle_qty
            display_total = vehicle_qty
        else:  # all
            if warehouse_qty == 0 and vehicle_qty == 0:
                continue
            display_warehouse = warehouse_qty
            display_vehicle = vehicle_qty
            display_total = warehouse_qty + vehicle_qty
        
        # Calculate values (using selling price)
        warehouse_value = display_warehouse * product.selling_price
        vehicle_value = display_vehicle * product.selling_price
        total_value = display_total * product.selling_price
        
        # Accumulate totals
        total_warehouse_qty += display_warehouse
        total_vehicle_qty += display_vehicle
        total_stock_qty += display_total
        total_warehouse_value += warehouse_value
        total_vehicle_value += vehicle_value
        total_stock_value += total_value
        
        stock_data.append({
            'product': product,
            'warehouse_qty': display_warehouse,
            'vehicle_qty': display_vehicle,
            'total_qty': display_total,
            'warehouse_value': warehouse_value,
            'vehicle_value': vehicle_value,
            'total_value': total_value,
        })
    
    # Sort by total quantity descending
    stock_data.sort(key=lambda x: x['total_qty'], reverse=True)
    
    # Get categories for filter
    categories = Category.objects.filter(is_active=True)
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory by Location"
        
        headers = ['Product', 'Category', 'Unit', 'Warehouse Qty', 'Vehicle Qty', 'Total Qty', 'Total Value (Rs)']
        ws.append(headers)
        for col in range(1, 8):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        for row in stock_data:
            ws.append([
                row['product'].name,
                row['product'].category.name if row['product'].category else 'Uncategorized',
                row['product'].unit,
                float(row['warehouse_qty']),
                float(row['vehicle_qty']),
                float(row['total_qty']),
                float(row['total_value']),
            ])
        
        # Add summary row
        ws.append([])
        ws.append(['TOTALS', '', '', float(total_warehouse_qty), float(total_vehicle_qty), float(total_stock_qty), float(total_stock_value)])
        for col in range(1, 8):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Inventory_by_Location_{today.strftime("%Y-%m-%d")}.xlsx"'
        wb.save(response)
        return response
    
    context = {
        'stock_data': stock_data,
        'location_filter': location_filter,
        'categories': categories,
        'selected_category': category_id,
        'total_warehouse_qty': total_warehouse_qty,
        'total_vehicle_qty': total_vehicle_qty,
        'total_stock_qty': total_stock_qty,
        'total_warehouse_value': total_warehouse_value,
        'total_vehicle_value': total_vehicle_value,
        'total_stock_value': total_stock_value,
        'today': today,
    }
    
    return render(request, 'core/inventory_by_location_report.html', context)


@login_required
@permission_required('view_reports')
def vehicle_loading_history_report(request):
    from datetime import date, datetime
    from decimal import Decimal
    from django.db.models import Sum
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from django.http import HttpResponse

    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    vehicle_id = request.GET.get('vehicle', '')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today

    # Safe query: handle empty table gracefully
    try:
        loads = VehicleLoad.objects.filter(
            loaded_at__date__gte=start_date_obj,
            loaded_at__date__lte=end_date_obj
        ).select_related('vehicle', 'product')
        
        if vehicle_id and vehicle_id.isdigit():
            loads = loads.filter(vehicle_id=int(vehicle_id))
        
        total_loads = loads.count()
        total_quantity = loads.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    except Exception as e:
        # If the table doesn't exist, we'll just return empty
        loads = []
        total_loads = 0
        total_quantity = Decimal('0')

    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicle Loading History"
        headers = ['Date', 'Vehicle', 'Product', 'Quantity', 'Notes']
        ws.append(headers)
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for load in loads:
            ws.append([
                load.loaded_at.strftime('%Y-%m-%d %H:%M'),
                load.vehicle.vehicle_number,
                load.product.name,
                float(load.quantity),
                load.notes or '',
            ])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Vehicle_Loading_History_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response

    vehicles = Vehicle.objects.filter(is_active=True)
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'loads': loads,
        'total_loads': total_loads,
        'total_quantity': total_quantity,
        'vehicles': vehicles,
        'selected_vehicle': vehicle_id,
        'today': today,
    }
    return render(request, 'core/vehicle_loading_history_report.html', context)


@login_required
@permission_required('view_reports')
def vehicle_transfer_history_report(request):
    today = date.today()
    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    source_vehicle_id = request.GET.get('source_vehicle', '')
    dest_vehicle_id = request.GET.get('dest_vehicle', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today
        end_date_obj = today
    
    transfers = StockTransfer.objects.filter(
        transfer_date__date__gte=start_date_obj,
        transfer_date__date__lte=end_date_obj
    ).select_related('source_vehicle', 'destination_vehicle', 'product', 'transferred_by')
    
    if source_vehicle_id and source_vehicle_id.isdigit():
        transfers = transfers.filter(source_vehicle_id=int(source_vehicle_id))
    if dest_vehicle_id and dest_vehicle_id.isdigit():
        transfers = transfers.filter(destination_vehicle_id=int(dest_vehicle_id))
    
    total_transfers = transfers.count()
    total_quantity = transfers.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicle Transfer History"
        headers = ['Date', 'From Vehicle', 'To Vehicle', 'Product', 'Quantity', 'Transferred By', 'Reason', 'Notes']
        ws.append(headers)
        for col in range(1, 9):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for t in transfers:
            ws.append([
                t.transfer_date.strftime('%Y-%m-%d %H:%M'),
                t.source_vehicle.vehicle_number,
                t.destination_vehicle.vehicle_number,
                t.product.name,
                float(t.quantity),
                t.transferred_by.username if t.transferred_by else '',
                t.get_reason_display(),
                t.notes or '',
            ])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Vehicle_Transfer_History_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    vehicles = Vehicle.objects.filter(is_active=True)
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'transfers': transfers,
        'total_transfers': total_transfers,
        'total_quantity': total_quantity,
        'vehicles': vehicles,
        'selected_source': source_vehicle_id,
        'selected_dest': dest_vehicle_id,
        'today': today,
    }
    return render(request, 'core/vehicle_transfer_history_report.html', context)


@login_required
@permission_required('view_reports')
def purchase_products_report(request):
    today = date.today()
    start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    supplier_id = request.GET.get('supplier', '')
    product_id = request.GET.get('product', '')
    category_id = request.GET.get('category', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today.replace(day=1)
        end_date_obj = today
    
    # Base queryset
    items = PurchaseItem.objects.filter(
        purchase__purchase_date__gte=start_date_obj,
        purchase__purchase_date__lte=end_date_obj
    ).select_related('purchase', 'purchase__supplier', 'product', 'product__category')
    
    if supplier_id and supplier_id.isdigit():
        items = items.filter(purchase__supplier_id=int(supplier_id))
    if product_id and product_id.isdigit():
        items = items.filter(product_id=int(product_id))
    if category_id and category_id.isdigit():
        items = items.filter(product__category_id=int(category_id))
    
    # Exclude FOC items? For this report we include all (both regular and FOC)
    # We'll add a column to indicate FOC status.
    
    total_qty = items.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    total_cost = items.aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_invoices = items.values('purchase_id').distinct().count()
    suppliers_used = items.values('purchase__supplier_id').distinct().count()
    
    # For Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Products History"
        headers = ['Date', 'Invoice No', 'Supplier', 'Product', 'Category', 'Qty', 'Unit Price', 'Total Cost', 'FOC?']
        ws.append(headers)
        for col in range(1, 10):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        for item in items:
            ws.append([
                item.purchase.purchase_date.strftime('%Y-%m-%d'),
                item.purchase.invoice_no,
                item.purchase.supplier.name,
                item.product.name,
                item.product.category.name if item.product.category else '',
                float(item.quantity),
                float(item.cost_price),
                float(item.total),
                'Yes' if item.is_foc else 'No',
            ])
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Purchase_Products_History_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    # Get filter options
    suppliers = Supplier.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True)
    categories = Category.objects.filter(is_active=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'items': items,
        'total_qty': total_qty,
        'total_cost': total_cost,
        'total_invoices': total_invoices,
        'suppliers_used': suppliers_used,
        'suppliers': suppliers,
        'products': products,
        'categories': categories,
        'selected_supplier': supplier_id,
        'selected_product': product_id,
        'selected_category': category_id,
    }
    return render(request, 'core/purchase_products_report.html', context)


@login_required
@permission_required('view_reports')
def purchase_foc_report(request):
    today = date.today()
    start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    supplier_id = request.GET.get('supplier', '')
    product_id = request.GET.get('product', '')
    category_id = request.GET.get('category', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today.replace(day=1)
        end_date_obj = today
    
    # Only FOC items
    items = PurchaseItem.objects.filter(
        is_foc=True,
        purchase__purchase_date__gte=start_date_obj,
        purchase__purchase_date__lte=end_date_obj
    ).select_related('purchase', 'purchase__supplier', 'product', 'product__category')
    
    if supplier_id and supplier_id.isdigit():
        items = items.filter(purchase__supplier_id=int(supplier_id))
    if product_id and product_id.isdigit():
        items = items.filter(product_id=int(product_id))
    if category_id and category_id.isdigit():
        items = items.filter(product__category_id=int(category_id))
    
    total_foc_qty = items.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    total_foc_value = items.aggregate(total=Sum('total'))['total'] or Decimal('0')
    suppliers_with_foc = items.values('purchase__supplier_id').distinct().count()
    unique_foc_products = items.values('product_id').distinct().count()
    
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "FOC Products Received"
        headers = ['Date', 'Invoice No', 'Supplier', 'Product', 'Category', 'Qty', 'Unit Value', 'Total FOC Value']
        ws.append(headers)
        for col in range(1, 9):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        for item in items:
            ws.append([
                item.purchase.purchase_date.strftime('%Y-%m-%d'),
                item.purchase.invoice_no,
                item.purchase.supplier.name,
                item.product.name,
                item.product.category.name if item.product.category else '',
                float(item.quantity),
                float(item.cost_price),
                float(item.total),
            ])
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="FOC_Products_Received_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    
    suppliers = Supplier.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True)
    categories = Category.objects.filter(is_active=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'items': items,
        'total_foc_qty': total_foc_qty,
        'total_foc_value': total_foc_value,
        'suppliers_with_foc': suppliers_with_foc,
        'unique_foc_products': unique_foc_products,
        'suppliers': suppliers,
        'products': products,
        'categories': categories,
        'selected_supplier': supplier_id,
        'selected_product': product_id,
        'selected_category': category_id,
    }
    return render(request, 'core/purchase_foc_report.html', context)


@login_required
@permission_required('view_reports')
def monthly_purchase_report(request):
    today = date.today()
    year = request.GET.get('year', str(today.year))
    supplier_id = request.GET.get('supplier', '')
    category_id = request.GET.get('category', '')
    
    try:
        year_int = int(year)
    except ValueError:
        year_int = today.year
    
    # Base queryset for purchases in that year
    purchases = Purchase.objects.filter(
        purchase_date__year=year_int,
        status__in=['RECEIVED', 'COMPLETED']  # Only completed/received purchases
    ).select_related('supplier')
    
    if supplier_id and supplier_id.isdigit():
        purchases = purchases.filter(supplier_id=int(supplier_id))
    
    # Category filter applies to items, so we need to filter through items
    if category_id and category_id.isdigit():
        purchases = purchases.filter(items__product__category_id=int(category_id)).distinct()
    
    # Group by month
    monthly_data = []
    for month in range(1, 13):
        month_purchases = purchases.filter(purchase_date__month=month)
        if month_purchases.exists():
            total_qty = PurchaseItem.objects.filter(purchase__in=month_purchases).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
            total_cost = PurchaseItem.objects.filter(purchase__in=month_purchases).aggregate(total=Sum('total'))['total'] or Decimal('0')
            foc_qty = PurchaseItem.objects.filter(purchase__in=month_purchases, is_foc=True).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
            foc_value = PurchaseItem.objects.filter(purchase__in=month_purchases, is_foc=True).aggregate(total=Sum('total'))['total'] or Decimal('0')
            invoices = month_purchases.count()
            suppliers = month_purchases.values('supplier_id').distinct().count()
            avg_cost_per_invoice = total_cost / invoices if invoices > 0 else Decimal('0')
            
            monthly_data.append({
                'month': month,
                'invoices': invoices,
                'total_qty': total_qty,
                'total_cost': total_cost,
                'foc_qty': foc_qty,
                'foc_value': foc_value,
                'suppliers': suppliers,
                'avg_cost_per_invoice': avg_cost_per_invoice,
            })
        else:
            monthly_data.append({
                'month': month,
                'invoices': 0,
                'total_qty': Decimal('0'),
                'total_cost': Decimal('0'),
                'foc_qty': Decimal('0'),
                'foc_value': Decimal('0'),
                'suppliers': 0,
                'avg_cost_per_invoice': Decimal('0'),
            })
    
    # Totals
    total_invoices = sum(m['invoices'] for m in monthly_data)
    total_cost_all = sum(m['total_cost'] for m in monthly_data)
    total_foc_value_all = sum(m['foc_value'] for m in monthly_data)
    
    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Purchase Report"
        headers = ['Month', 'Invoices', 'Total Qty', 'Total Cost', 'FOC Qty', 'FOC Value', 'Suppliers', 'Avg Cost/Invoice']
        ws.append(headers)
        for col in range(1, 9):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        for data in monthly_data:
            ws.append([
                f"{year_int}-{data['month']:02d}",
                data['invoices'],
                float(data['total_qty']),
                float(data['total_cost']),
                float(data['foc_qty']),
                float(data['foc_value']),
                data['suppliers'],
                float(data['avg_cost_per_invoice']),
            ])
        
        ws.append([])
        ws.append(['TOTAL', total_invoices, '', float(total_cost_all), '', float(total_foc_value_all), '', ''])
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Monthly_Purchase_Report_{year_int}.xlsx"'
        wb.save(response)
        return response
    
    suppliers = Supplier.objects.filter(is_active=True)
    categories = Category.objects.filter(is_active=True)
    
    # Get list of years available for filter
    years = Purchase.objects.dates('purchase_date', 'year').values_list('purchase_date__year', flat=True).distinct()
    years = sorted(list(years), reverse=True)
    
    context = {
        'year': year_int,
        'years': years,
        'suppliers': suppliers,
        'categories': categories,
        'selected_supplier': supplier_id,
        'selected_category': category_id,
        'monthly_data': monthly_data,
        'total_invoices': total_invoices,
        'total_cost_all': total_cost_all,
        'total_foc_value_all': total_foc_value_all,
    }
    return render(request, 'core/monthly_purchase_report.html', context)


@login_required
@permission_required('view_reports')
def sales_credit_report(request):
    try:
        from datetime import datetime, date
        from decimal import Decimal
        from django.db.models import Sum, Q
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from django.http import HttpResponse

        today = date.today()
        start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
        end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
        vehicle_id = request.GET.get('vehicle', '')

        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = today.replace(day=1)
            end_date_obj = today

        # Base queryset: all credit payments
        credit_bill_ids = Payment.objects.filter(
            type='Credit',
            bill__date__gte=start_date_obj,
            bill__date__lte=end_date_obj,
            bill__status='COMPLETED'
        ).values_list('bill_id', flat=True).distinct()

        bills = SalesBill.objects.filter(id__in=credit_bill_ids).select_related('vehicle', 'rep')
        
        if vehicle_id and vehicle_id.isdigit():
            bills = bills.filter(vehicle_id=int(vehicle_id))

        # Get selected vehicle object for display
        selected_vehicle_obj = None
        if vehicle_id and vehicle_id.isdigit():
            try:
                selected_vehicle_obj = Vehicle.objects.get(id=vehicle_id)
            except Vehicle.DoesNotExist:
                pass

        # Process bills: compute paid, outstanding, collection status
        bill_data = []
        total_credit_sales = Decimal('0')
        total_outstanding = Decimal('0')
        total_paid = Decimal('0')
        customers_set = set()

        for bill in bills:
            paid_amount = bill.payments.exclude(type='Credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            outstanding = bill.net_total - paid_amount
            total_credit_sales += bill.net_total
            total_outstanding += outstanding
            total_paid += paid_amount
            customers_set.add(bill.shop_name or bill.shop_code or 'N/A')

            # Collection status
            collection = CreditCollection.objects.filter(sales_bill=bill).first()
            collection_status = collection.status if collection else 'PENDING'
            collection_rep = collection.rep.name if collection and collection.rep else None

            bill_data.append({
                'bill': bill,
                'paid': paid_amount,
                'outstanding': outstanding,
                'collection_status': collection_status,
                'collection_rep': collection_rep,
            })

        total_customers = len(customers_set)

        # Group by vehicle
        vehicle_summary = {}
        for item in bill_data:
            vehicle = item['bill'].vehicle
            v_key = vehicle.id if vehicle else None
            if v_key not in vehicle_summary:
                vehicle_summary[v_key] = {
                    'vehicle': vehicle,
                    'total_bills': 0,
                    'total_sales': Decimal('0'),
                    'total_paid': Decimal('0'),
                    'total_outstanding': Decimal('0'),
                    'bills': [],
                }
            vehicle_summary[v_key]['total_bills'] += 1
            vehicle_summary[v_key]['total_sales'] += item['bill'].net_total
            vehicle_summary[v_key]['total_paid'] += item['paid']
            vehicle_summary[v_key]['total_outstanding'] += item['outstanding']
            vehicle_summary[v_key]['bills'].append(item)

        # Convert to list for template, sort by sales descending
        vehicle_summary_list = sorted(vehicle_summary.values(), key=lambda x: x['total_sales'], reverse=True)

        # If a specific vehicle is selected, show detailed bills instead of summary
        show_detail = bool(vehicle_id and vehicle_id.isdigit())

        # Excel export
        if request.GET.get('export') == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Credit Report"

            if show_detail:
                headers = ['Invoice', 'Date', 'Customer', 'Vehicle', 'Total Amount', 'Paid', 'Outstanding', 'Collection Status']
                ws.append(headers)
                for col in range(1, 9):
                    ws.cell(row=1, column=col).font = Font(bold=True)
                for item in bill_data:
                    ws.append([
                        item['bill'].invoice_no,
                        item['bill'].date.strftime('%Y-%m-%d'),
                        item['bill'].shop_name or 'N/A',
                        item['bill'].vehicle.vehicle_number if item['bill'].vehicle else 'N/A',
                        float(item['bill'].net_total),
                        float(item['paid']),
                        float(item['outstanding']),
                        item['collection_status'],
                    ])
            else:
                headers = ['Vehicle', 'Total Bills', 'Total Sales', 'Total Paid', 'Total Outstanding', 'Collection Rate %']
                ws.append(headers)
                for col in range(1, 7):
                    ws.cell(row=1, column=col).font = Font(bold=True)
                vehicle_summary_list = []
                for key, data in vehicle_summary.items():
                    rate = (data['total_paid'] / data['total_sales'] * 100) if data['total_sales'] > 0 else 0
                    vehicle_summary_list.append({
                        'vehicle': data['vehicle'],
                        'total_bills': data['total_bills'],
                        'total_sales': data['total_sales'],
                        'total_paid': data['total_paid'],
                        'total_outstanding': data['total_outstanding'],
                        'rate': rate,
                    })

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Sales_Credit_Report_{start_date}_to_{end_date}.xlsx"'
            wb.save(response)
            return response

        # Vehicles for filter
        vehicles = Vehicle.objects.filter(is_active=True)

        context = {
            'start_date': start_date,
            'end_date': end_date,
            'start_date_obj': start_date_obj,
            'end_date_obj': end_date_obj,
            'vehicles': vehicles,
            'selected_vehicle': vehicle_id,
            'selected_vehicle_obj': selected_vehicle_obj,
            'show_detail': show_detail,
            'bill_data': bill_data,
            'vehicle_summary_list': vehicle_summary_list,
            'total_credit_sales': total_credit_sales,
            'total_outstanding': total_outstanding,
            'total_paid': total_paid,
            'total_customers': total_customers,
            'total_bills': len(bill_data),
            'today': today,
        }
        return render(request, 'core/sales_credit_report.html', context)
    
    except Exception as e:
        import traceback
        print(f"ERROR in sales_credit_report: {e}")
        print(traceback.format_exc())
        # Return a simple error page with the error details (for debugging)
        return HttpResponse(f"<h1>Error</h1><p>{e}</p><pre>{traceback.format_exc()}</pre>", status=500)


@login_required
@permission_required('view_reports')
def credit_purchase_report(request):
    today = date.today()
    start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    supplier_id = request.GET.get('supplier', '')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today.replace(day=1)
        end_date_obj = today

    # Get all purchases with payment status = CREDIT (not fully paid)
    purchases = Purchase.objects.filter(
        purchase_date__gte=start_date_obj,
        purchase_date__lte=end_date_obj,
        payment_status='CREDIT',
        status__in=['RECEIVED', 'COMPLETED']
    ).select_related('supplier', 'rep')

    if supplier_id and supplier_id.isdigit():
        purchases = purchases.filter(supplier_id=int(supplier_id))

    # Calculate outstanding for each purchase
    purchase_data = []
    total_credit_amount = Decimal('0')
    total_paid_amount = Decimal('0')
    total_outstanding = Decimal('0')

    for purchase in purchases:
        outstanding = purchase.total - purchase.paid_amount
        purchase_data.append({
            'purchase': purchase,
            'outstanding': outstanding,
        })
        total_credit_amount += purchase.total
        total_paid_amount += purchase.paid_amount
        total_outstanding += outstanding

    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Credit Purchase Report"

        headers = ['Date', 'Invoice No', 'Supplier', 'PO Number', 'Total Amount', 'Paid', 'Outstanding', 'Due Date']
        ws.append(headers)
        for col in range(1, 9):
            ws.cell(row=1, column=col).font = Font(bold=True)

        for item in purchase_data:
            p = item['purchase']
            ws.append([
                p.purchase_date.strftime('%Y-%m-%d'),
                p.invoice_no,
                p.supplier.name,
                p.po_number or '',
                float(p.total),
                float(p.paid_amount),
                float(item['outstanding']),
                p.due_date.strftime('%Y-%m-%d') if p.due_date else '',
            ])

        # Add total row
        ws.append([])
        ws.append(['TOTALS', '', '', '', float(total_credit_amount), float(total_paid_amount), float(total_outstanding), ''])
        for col in range(1, 9):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Credit_Purchase_Report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response

    # Suppliers for filter
    suppliers = Supplier.objects.filter(is_active=True)

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'suppliers': suppliers,
        'selected_supplier': supplier_id,
        'purchase_data': purchase_data,
        'total_credit_amount': total_credit_amount,
        'total_paid_amount': total_paid_amount,
        'total_outstanding': total_outstanding,
        'total_purchases': len(purchase_data),
        'today': today,
    }
    return render(request, 'core/credit_purchase_report.html', context)


@login_required
@permission_required('view_reports')
def expense_by_vehicle_summary_report(request):

    today = date.today()
    start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    category_filter = request.GET.get('category', '')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today.replace(day=1)
        end_date_obj = today

    # Base queryset
    expenses = Expense.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )

    if category_filter:
        expenses = expenses.filter(category=category_filter)

    # Get all vehicles with expenses
    vehicle_ids = expenses.values_list('vehicle_id', flat=True).distinct()
    vehicles = Vehicle.objects.filter(id__in=vehicle_ids)

    # Build vehicle-wise data
    vehicle_data = []
    total_expenses = Decimal('0')
    expense_categories = set()

    for vehicle in vehicles:
        veh_expenses = expenses.filter(vehicle=vehicle)
        veh_total = veh_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_expenses += veh_total

        # Category breakdown for this vehicle
        category_breakdown = {}
        for cat in Expense.CATEGORY_CHOICES:
            cat_key = cat[0]
            cat_total = veh_expenses.filter(category=cat_key).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            if cat_total > 0:
                category_breakdown[cat_key] = cat_total
                expense_categories.add(cat_key)

        vehicle_data.append({
            'vehicle': vehicle,
            'total': veh_total,
            'categories': category_breakdown,
        })

    # Include unassigned expenses (vehicle=None)
    unassigned_expenses = expenses.filter(vehicle__isnull=True)
    if unassigned_expenses.exists():
        veh_total = unassigned_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_expenses += veh_total
        category_breakdown = {}
        for cat in Expense.CATEGORY_CHOICES:
            cat_key = cat[0]
            cat_total = unassigned_expenses.filter(category=cat_key).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            if cat_total > 0:
                category_breakdown[cat_key] = cat_total
                expense_categories.add(cat_key)
        vehicle_data.append({
            'vehicle': None,
            'total': veh_total,
            'categories': category_breakdown,
        })

    # Sort by total descending
    vehicle_data.sort(key=lambda x: x['total'], reverse=True)

    # Build category totals for the footer
    category_totals = {}
    for cat in expense_categories:
        cat_total = expenses.filter(category=cat).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        category_totals[cat] = cat_total

    # Prepare a list of categories for the template header
    sorted_categories = sorted(expense_categories)
    # Build a list of category display names
    category_display_names = {cat[0]: cat[1] for cat in Expense.CATEGORY_CHOICES}

    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense by Vehicle Summary"

        headers = ['Vehicle', 'Driver']
        for cat in sorted_categories:
            headers.append(category_display_names.get(cat, cat))
        headers.append('Total')

        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        for item in vehicle_data:
            vehicle_name = item['vehicle'].vehicle_number if item['vehicle'] else 'Unassigned'
            driver_name = item['vehicle'].driver_name if item['vehicle'] else ''
            row = [vehicle_name, driver_name]
            for cat in sorted_categories:
                row.append(float(item['categories'].get(cat, 0)))
            row.append(float(item['total']))
            ws.append(row)

        # Grand total row
        total_row = ['GRAND TOTAL', '']
        for cat in sorted_categories:
            total_row.append(float(category_totals.get(cat, 0)))
        total_row.append(float(total_expenses))
        ws.append(total_row)
        for col in range(1, len(total_row) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Expense_by_Vehicle_Summary_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response

    # For the template, we'll pass the data ready to display without custom filters
    # We'll build a list of rows with category values in the same order as sorted_categories
    vehicle_rows = []
    for item in vehicle_data:
        cat_values = []
        for cat in sorted_categories:
            cat_values.append(item['categories'].get(cat, 0))
        vehicle_rows.append({
            'vehicle': item['vehicle'],
            'total': item['total'],
            'cat_values': cat_values,
        })

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'vehicle_rows': vehicle_rows,
        'category_names': [category_display_names.get(cat, cat) for cat in sorted_categories],
        'category_totals_values': [category_totals.get(cat, 0) for cat in sorted_categories],
        'total_expenses': total_expenses,
        'total_vehicles': len(vehicle_data),
        'categories': Expense.CATEGORY_CHOICES,
        'selected_category': category_filter,
        'today': today,
    }
    return render(request, 'core/expense_by_vehicle_summary_report.html', context)


@login_required
@permission_required('view_reports')
def expense_by_employee_summary_report(request):
    today = date.today()
    start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    category_filter = request.GET.get('category', '')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date_obj = today.replace(day=1)
        end_date_obj = today

    # Base queryset
    expenses = Expense.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )

    if category_filter:
        expenses = expenses.filter(category=category_filter)

    # Get all employees with expenses
    employee_ids = expenses.values_list('employee_id', flat=True).distinct()
    employees = Employee.objects.filter(id__in=employee_ids)

    # Build employee-wise data
    employee_data = []
    total_expenses = Decimal('0')
    expense_categories = set()

    for employee in employees:
        emp_expenses = expenses.filter(employee=employee)
        emp_total = emp_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_expenses += emp_total

        # Category breakdown for this employee
        category_breakdown = {}
        for cat in Expense.CATEGORY_CHOICES:
            cat_key = cat[0]
            cat_total = emp_expenses.filter(category=cat_key).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            if cat_total > 0:
                category_breakdown[cat_key] = cat_total
                expense_categories.add(cat_key)

        employee_data.append({
            'employee': employee,
            'total': emp_total,
            'categories': category_breakdown,
        })

    # Also include expenses with no employee (unassigned)
    unassigned_expenses = expenses.filter(employee__isnull=True)
    if unassigned_expenses.exists():
        emp_total = unassigned_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_expenses += emp_total
        category_breakdown = {}
        for cat in Expense.CATEGORY_CHOICES:
            cat_key = cat[0]
            cat_total = unassigned_expenses.filter(category=cat_key).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            if cat_total > 0:
                category_breakdown[cat_key] = cat_total
                expense_categories.add(cat_key)
        employee_data.append({
            'employee': None,  # Unassigned
            'total': emp_total,
            'categories': category_breakdown,
        })

    # Sort by total descending
    employee_data.sort(key=lambda x: x['total'], reverse=True)

    # Category totals across all employees
    category_totals = {}
    for cat in expense_categories:
        cat_total = expenses.filter(category=cat).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        category_totals[cat] = cat_total

    # Excel export
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense by Employee Summary"

        # Headers
        headers = ['Employee', 'Position']
        sorted_categories = sorted(expense_categories)
        for cat in sorted_categories:
            display_name = dict(Expense.CATEGORY_CHOICES).get(cat, cat)
            headers.append(display_name)
        headers.append('Total')

        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        # Data rows
        for item in employee_data:
            emp_name = item['employee'].name if item['employee'] else 'Unassigned'
            emp_position = item['employee'].get_position_display() if item['employee'] else ''
            row = [emp_name, emp_position]
            for cat in sorted_categories:
                row.append(float(item['categories'].get(cat, 0)))
            row.append(float(item['total']))
            ws.append(row)

        # Grand total row
        total_row = ['GRAND TOTAL', '']
        for cat in sorted_categories:
            total_row.append(float(category_totals.get(cat, 0)))
        total_row.append(float(total_expenses))
        ws.append(total_row)
        for col in range(1, len(total_row) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Expense_by_Employee_Summary_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response

    # All categories for filter dropdown
    categories = Expense.CATEGORY_CHOICES

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_obj': start_date_obj,
        'end_date_obj': end_date_obj,
        'employee_data': employee_data,
        'category_totals': category_totals,
        'total_expenses': total_expenses,
        'expense_categories': sorted(expense_categories),
        'categories': categories,
        'selected_category': category_filter,
        'total_employees': len(employee_data),
        'today': today,
    }
    return render(request, 'core/expense_by_employee_summary_report.html', context)
